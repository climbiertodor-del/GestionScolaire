import tkinter as tk
from tkinter import messagebox, filedialog, ttk
import sqlite3

try:
    from openpyxl import load_workbook, Workbook
except ImportError:
    load_workbook = None
    Workbook = None
from pathlib import Path
import shutil


# ============================================================
# GESTIONSCOLAIRE
# APPLICATION WINDOWS DESKTOP LOCALE
# ============================================================

BASE_PROJET = Path(__file__).resolve().parent

DOSSIER_DATA = BASE_PROJET / "data"
DOSSIER_DATA.mkdir(exist_ok=True)

BASE_DONNEES = DOSSIER_DATA / "gestion_scolaire.db"

DOSSIER_LOGO = DOSSIER_DATA / "logo"
DOSSIER_LOGO.mkdir(exist_ok=True)


# ============================================================
# COULEURS
# ============================================================

MENU = "#243447"
MENU_HOVER = "#34495E"
MENU_ACTIF = "#167D7F"

TURQUOISE = "#167D7F"
TURQUOISE_VIF = "#5EEAD4"

FOND = "#EEF2F5"
CARTE = "#F8FAFC"
BLANC = "#FFFFFF"

BORDURE = "#D9E2E8"

BLEU = "#3B82F6"
VERT = "#22A06B"
ORANGE = "#E9A23B"
ROUGE = "#D9534F"

TEXTE = "#263746"
TEXTE_SECONDAIRE = "#718096"

INFORMATIONS_FOND = "#E8F3F1"
ALERTES_FOND = "#FFF4E3"


# ============================================================
# BASE DE DONNÉES
# ============================================================

def connexion():
    db = sqlite3.connect(BASE_DONNEES)
    db.row_factory = sqlite3.Row
    return db


def initialiser_base():

    db = connexion()
    curseur = db.cursor()

    # --------------------------------------------------------
    # ÉTABLISSEMENT
    # --------------------------------------------------------

    curseur.execute("""
        CREATE TABLE IF NOT EXISTS etablissement (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nom TEXT NOT NULL,
            type_etablissement TEXT,
            adresse TEXT,
            telephone TEXT,
            email TEXT,
            directeur TEXT,
            secretaire TEXT,
            logo TEXT,
            informations TEXT,
            localite TEXT,
            dre_inspection TEXT,
            devise TEXT
        )
    """)

    # --------------------------------------------------------
    # ANNÉES SCOLAIRES
    # --------------------------------------------------------

    curseur.execute("""
        CREATE TABLE IF NOT EXISTS annees_scolaires (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            libelle TEXT NOT NULL UNIQUE,
            date_debut TEXT,
            date_fin TEXT,
            active INTEGER DEFAULT 0,
            cloturee INTEGER DEFAULT 0
        )
    """)

    # --------------------------------------------------------
    # STRUCTURE DES CLASSES ET DES ÉLÈVES
    # --------------------------------------------------------

    curseur.execute("""
        CREATE TABLE IF NOT EXISTS classes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            code TEXT NOT NULL UNIQUE,
            niveau TEXT NOT NULL,
            libelle TEXT NOT NULL,
            serie TEXT,
            effectif_max INTEGER,
            actif INTEGER DEFAULT 1
        )
    """)

    curseur.execute("""
        CREATE TABLE IF NOT EXISTS eleves (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            matricule TEXT NOT NULL UNIQUE,
            nom TEXT NOT NULL,
            prenom TEXT NOT NULL,
            sexe TEXT NOT NULL,
            date_naissance TEXT,
            lieu_naissance TEXT,
            nationalite TEXT,
            statut_entree TEXT NOT NULL DEFAULT 'Nouveau',
            situation_actuelle TEXT NOT NULL DEFAULT 'Actif',
            nom_parent TEXT,
            telephone_parent TEXT,
            observation TEXT,
            date_creation TEXT DEFAULT CURRENT_TIMESTAMP
        )
    """)

    curseur.execute("""
        CREATE TABLE IF NOT EXISTS inscriptions_eleves (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            eleve_id INTEGER NOT NULL,
            annee_id INTEGER NOT NULL,
            classe_id INTEGER,
            statut TEXT NOT NULL DEFAULT 'Inscrit',
            date_inscription TEXT DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(eleve_id, annee_id),
            FOREIGN KEY(eleve_id) REFERENCES eleves(id) ON DELETE CASCADE,
            FOREIGN KEY(annee_id) REFERENCES annees_scolaires(id),
            FOREIGN KEY(classe_id) REFERENCES classes(id)
        )
    """)

    # --------------------------------------------------------
    # COMPATIBILITÉ ANCIENNE BASE
    # --------------------------------------------------------

    curseur.execute("PRAGMA table_info(etablissement)")

    colonnes = {
        ligne["name"]
        for ligne in curseur.fetchall()
    }

    nouvelles_colonnes = {
        "localite": "TEXT",
        "dre_inspection": "TEXT",
        "devise": "TEXT"
    }

    for colonne, type_colonne in nouvelles_colonnes.items():

        if colonne not in colonnes:

            curseur.execute(
                f"""
                ALTER TABLE etablissement
                ADD COLUMN {colonne} {type_colonne}
                """
            )

    db.commit()
    db.close()


# ============================================================
# LECTURE DES DONNÉES
# ============================================================

def configuration_existe():

    db = connexion()
    curseur = db.cursor()

    curseur.execute("""
        SELECT COUNT(*) AS total
        FROM etablissement
    """)

    resultat = curseur.fetchone()

    db.close()

    return resultat["total"] > 0


def obtenir_etablissement():

    db = connexion()
    curseur = db.cursor()

    curseur.execute("""
        SELECT *
        FROM etablissement
        ORDER BY id
        LIMIT 1
    """)

    resultat = curseur.fetchone()

    db.close()

    return resultat


def obtenir_annee_active():

    db = connexion()
    curseur = db.cursor()

    curseur.execute("""
        SELECT *
        FROM annees_scolaires
        WHERE active = 1
        LIMIT 1
    """)

    resultat = curseur.fetchone()

    db.close()

    return resultat


# ============================================================
# ENREGISTRER L'ÉCOLE
# ============================================================

def enregistrer_ecole(
    nom,
    type_etablissement,
    adresse,
    localite,
    telephone,
    email,
    dre_inspection,
    directeur,
    secretaire,
    devise,
    logo,
    annee
):

    db = connexion()
    curseur = db.cursor()

    curseur.execute("""
        SELECT id
        FROM etablissement
        ORDER BY id
        LIMIT 1
    """)

    ecole_existante = curseur.fetchone()

    if ecole_existante:

        curseur.execute("""
            UPDATE etablissement
            SET
                nom = ?,
                type_etablissement = ?,
                adresse = ?,
                localite = ?,
                telephone = ?,
                email = ?,
                dre_inspection = ?,
                directeur = ?,
                secretaire = ?,
                devise = ?,
                logo = ?
            WHERE id = ?
        """, (
            nom,
            type_etablissement,
            adresse,
            localite,
            telephone,
            email,
            dre_inspection,
            directeur,
            secretaire,
            devise,
            logo,
            ecole_existante["id"]
        ))

    else:

        curseur.execute("""
            INSERT INTO etablissement (
                nom,
                type_etablissement,
                adresse,
                localite,
                telephone,
                email,
                directeur,
                secretaire,
                logo,
                informations,
                dre_inspection,
                devise
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            nom,
            type_etablissement,
            adresse,
            localite,
            telephone,
            email,
            directeur,
            secretaire,
            logo,
            "",
            dre_inspection,
            devise
        ))

    # --------------------------------------------------------
    # ANNÉE SCOLAIRE
    # --------------------------------------------------------

    curseur.execute("""
        UPDATE annees_scolaires
        SET active = 0
    """)

    curseur.execute("""
        INSERT OR IGNORE INTO annees_scolaires (
            libelle,
            active,
            cloturee
        )
        VALUES (?, 0, 0)
    """, (annee,))

    curseur.execute("""
        UPDATE annees_scolaires
        SET active = 1
        WHERE libelle = ?
    """, (annee,))

    db.commit()
    db.close()


# ============================================================
# APPLICATION
# ============================================================

class GestionScolaire(tk.Tk):

    def __init__(self):

        super().__init__()

        self.title("GestionScolaire")

        self.geometry("1280x760")

        self.minsize(1100, 650)

        self.configure(bg=FOND)

        self.etablissement = None
        self.annee_active = None

        self.menu_buttons = {}

        self.logo_selectionne = ""

        initialiser_base()

        # ----------------------------------------------------
        # PREMIER DÉMARRAGE
        # ----------------------------------------------------

        if configuration_existe():

            self.charger_donnees()

            self.creer_interface()

            self.afficher_accueil()

        else:

            self.afficher_premiere_configuration()


    # ========================================================
    # CHARGER LES DONNÉES
    # ========================================================

    def charger_donnees(self):

        self.etablissement = obtenir_etablissement()

        self.annee_active = obtenir_annee_active()


    # ========================================================
    # PREMIÈRE CONFIGURATION
    # ========================================================

    def afficher_premiere_configuration(self):

        for widget in self.winfo_children():
            widget.destroy()

        self.title(
            "GestionScolaire - Configuration de l'école"
        )

        principal = tk.Frame(
            self,
            bg=FOND
        )

        principal.pack(
            fill="both",
            expand=True
        )

        # ----------------------------------------------------
        # ENTÊTE
        # ----------------------------------------------------

        entete = tk.Frame(
            principal,
            bg=TURQUOISE,
            height=105
        )

        entete.pack(fill="x")

        entete.pack_propagate(False)

        tk.Label(
            entete,
            text="GestionScolaire",
            bg=TURQUOISE,
            fg=BLANC,
            font=("Segoe UI", 24, "bold")
        ).pack(
            pady=(18, 0)
        )

        tk.Label(
            entete,
            text="CONFIGURATION DE L'ÉCOLE",
            bg=TURQUOISE,
            fg="#D5F5F2",
            font=("Segoe UI", 9, "bold")
        ).pack()

        # ----------------------------------------------------
        # ZONE
        # ----------------------------------------------------

        zone = tk.Frame(
            principal,
            bg=FOND
        )

        zone.pack(
            fill="both",
            expand=True,
            padx=70,
            pady=25
        )

        tk.Label(
            zone,
            text="Bienvenue dans GestionScolaire",
            bg=FOND,
            fg=TEXTE,
            font=("Segoe UI", 21, "bold")
        ).pack(anchor="w")

        tk.Label(
            zone,
            text=(
                "Saisissez les informations de votre établissement "
                "pour commencer."
            ),
            bg=FOND,
            fg=TEXTE_SECONDAIRE,
            font=("Segoe UI", 10)
        ).pack(
            anchor="w",
            pady=(4, 15)
        )

        carte = tk.Frame(
            zone,
            bg=CARTE,
            highlightbackground=BORDURE,
            highlightthickness=1
        )

        carte.pack(
            fill="both",
            expand=True
        )

        contenu = tk.Frame(
            carte,
            bg=CARTE
        )

        contenu.pack(
            fill="both",
            expand=True,
            padx=30,
            pady=20
        )

        # ----------------------------------------------------
        # NOM
        # ----------------------------------------------------

        self.champ_nom = self.creer_champ(
            contenu,
            "Nom de l'établissement *"
        )

        # ----------------------------------------------------
        # TYPE
        # ----------------------------------------------------

        tk.Label(
            contenu,
            text="Type d'établissement",
            bg=CARTE,
            fg=TEXTE,
            font=("Segoe UI", 9, "bold")
        ).pack(
            anchor="w",
            pady=(10, 4)
        )

        self.type_var = tk.StringVar(
            value="CEG"
        )

        menu_type = tk.OptionMenu(
            contenu,
            self.type_var,
            "CEG",
            "Collège",
            "Lycée",
            "Collège-Lycée",
            "École primaire",
            "Établissement privé",
            "Autre"
        )

        menu_type.configure(
            bg=BLANC,
            fg=TEXTE,
            activebackground=BLANC,
            activeforeground=TEXTE,
            relief="solid",
            bd=1,
            highlightthickness=0,
            font=("Segoe UI", 10)
        )

        menu_type.pack(
            fill="x"
        )

        # ----------------------------------------------------
        # LIGNE 1
        # ----------------------------------------------------

        ligne1 = tk.Frame(
            contenu,
            bg=CARTE
        )

        ligne1.pack(
            fill="x",
            pady=(10, 0)
        )

        gauche1 = tk.Frame(
            ligne1,
            bg=CARTE
        )

        gauche1.pack(
            side="left",
            fill="x",
            expand=True,
            padx=(0, 8)
        )

        droite1 = tk.Frame(
            ligne1,
            bg=CARTE
        )

        droite1.pack(
            side="left",
            fill="x",
            expand=True,
            padx=(8, 0)
        )

        self.champ_adresse = self.creer_champ(
            gauche1,
            "Adresse"
        )

        self.champ_localite = self.creer_champ(
            droite1,
            "Localité"
        )

        # ----------------------------------------------------
        # LIGNE 2
        # ----------------------------------------------------

        ligne2 = tk.Frame(
            contenu,
            bg=CARTE
        )

        ligne2.pack(
            fill="x",
            pady=(10, 0)
        )

        gauche2 = tk.Frame(
            ligne2,
            bg=CARTE
        )

        gauche2.pack(
            side="left",
            fill="x",
            expand=True,
            padx=(0, 8)
        )

        droite2 = tk.Frame(
            ligne2,
            bg=CARTE
        )

        droite2.pack(
            side="left",
            fill="x",
            expand=True,
            padx=(8, 0)
        )

        self.champ_telephone = self.creer_champ(
            gauche2,
            "Téléphone"
        )

        self.champ_email = self.creer_champ(
            droite2,
            "E-mail"
        )

        # ----------------------------------------------------
        # LIGNE 3
        # ----------------------------------------------------

        ligne3 = tk.Frame(
            contenu,
            bg=CARTE
        )

        ligne3.pack(
            fill="x",
            pady=(10, 0)
        )

        gauche3 = tk.Frame(
            ligne3,
            bg=CARTE
        )

        gauche3.pack(
            side="left",
            fill="x",
            expand=True,
            padx=(0, 8)
        )

        droite3 = tk.Frame(
            ligne3,
            bg=CARTE
        )

        droite3.pack(
            side="left",
            fill="x",
            expand=True,
            padx=(8, 0)
        )

        self.champ_directeur = self.creer_champ(
            gauche3,
            "Directeur / Chef d'établissement"
        )

        self.champ_secretaire = self.creer_champ(
            droite3,
            "Secrétaire"
        )

        # ----------------------------------------------------
        # LIGNE 4
        # ----------------------------------------------------

        ligne4 = tk.Frame(
            contenu,
            bg=CARTE
        )

        ligne4.pack(
            fill="x",
            pady=(10, 0)
        )

        gauche4 = tk.Frame(
            ligne4,
            bg=CARTE
        )

        gauche4.pack(
            side="left",
            fill="x",
            expand=True,
            padx=(0, 8)
        )

        droite4 = tk.Frame(
            ligne4,
            bg=CARTE
        )

        droite4.pack(
            side="left",
            fill="x",
            expand=True,
            padx=(8, 0)
        )

        self.champ_dre = self.creer_champ(
            gauche4,
            "DRE / Inspection"
        )

        self.champ_annee = self.creer_champ(
            droite4,
            "Année scolaire *"
        )

        # ----------------------------------------------------
        # DEVISE
        # ----------------------------------------------------

        self.champ_devise = self.creer_champ(
            contenu,
            "Devise / Slogan de l'établissement"
        )

        # ----------------------------------------------------
        # LOGO
        # ----------------------------------------------------

        bloc_logo = tk.Frame(
            contenu,
            bg=CARTE
        )

        bloc_logo.pack(
            fill="x",
            pady=(10, 0)
        )

        tk.Label(
            bloc_logo,
            text="Logo de l'établissement",
            bg=CARTE,
            fg=TEXTE,
            font=("Segoe UI", 9, "bold")
        ).pack(
            anchor="w",
            pady=(0, 4)
        )

        ligne_logo = tk.Frame(
            bloc_logo,
            bg=CARTE
        )

        ligne_logo.pack(
            fill="x"
        )

        self.logo_premiere_config = tk.StringVar(
            value="Aucun logo sélectionné"
        )

        tk.Label(
            ligne_logo,
            textvariable=self.logo_premiere_config,
            bg=BLANC,
            fg=TEXTE_SECONDAIRE,
            anchor="w",
            relief="solid",
            bd=1,
            padx=10,
            font=("Segoe UI", 9)
        ).pack(
            side="left",
            fill="x",
            expand=True,
            ipady=6
        )

        tk.Button(
            ligne_logo,
            text="CHOISIR LE LOGO",
            command=self.choisir_logo_premiere_config,
            bg=BLEU,
            fg=BLANC,
            activebackground="#2563EB",
            activeforeground=BLANC,
            relief="flat",
            bd=0,
            cursor="hand2",
            font=("Segoe UI", 9, "bold"),
            padx=15,
            pady=7
        ).pack(
            side="left",
            padx=(8, 0)
        )

        # ----------------------------------------------------
        # BOUTON
        # ----------------------------------------------------

        tk.Button(
            contenu,
            text="ENREGISTRER ET CONTINUER",
            command=self.valider_configuration,
            bg=TURQUOISE,
            fg=BLANC,
            activebackground="#115E59",
            activeforeground=BLANC,
            relief="flat",
            bd=0,
            cursor="hand2",
            font=("Segoe UI", 10, "bold"),
            padx=25,
            pady=11
        ).pack(
            anchor="e",
            pady=(15, 0)
        )

        self.champ_nom.focus_set()


    # ========================================================
    # CHAMP SIMPLE
    # ========================================================

    def creer_champ(
        self,
        parent,
        titre,
        valeur=""
    ):

        tk.Label(
            parent,
            text=titre,
            bg=CARTE,
            fg=TEXTE,
            font=("Segoe UI", 9, "bold")
        ).pack(
            anchor="w",
            pady=(0, 4)
        )

        champ = tk.Entry(
            parent,
            bg=BLANC,
            fg=TEXTE,
            insertbackground=TEXTE,
            relief="solid",
            bd=1,
            font=("Segoe UI", 10)
        )

        if valeur:
            champ.insert(0, valeur)

        champ.pack(
            fill="x",
            ipady=7
        )

        return champ


    # ========================================================
    # CHOISIR LOGO PREMIÈRE CONFIGURATION
    # ========================================================

    def choisir_logo_premiere_config(self):

        fichier = filedialog.askopenfilename(
            title="Choisir le logo de l'établissement",
            filetypes=[
                (
                    "Images",
                    "*.png *.jpg *.jpeg *.gif *.bmp"
                ),
                ("PNG", "*.png"),
                ("JPEG", "*.jpg *.jpeg"),
                ("Tous les fichiers", "*.*")
            ]
        )

        if not fichier:
            return

        try:

            source = Path(fichier)

            destination = DOSSIER_LOGO / (
                "logo_etablissement"
                + source.suffix.lower()
            )

            shutil.copy2(
                source,
                destination
            )

            self.logo_selectionne = str(destination)

            self.logo_premiere_config.set(
                source.name
            )

        except Exception as erreur:

            messagebox.showerror(
                "Erreur",
                f"Impossible de copier le logo.\n\n{erreur}"
            )


    # ========================================================
    # VALIDATION CONFIGURATION
    # ========================================================

    def valider_configuration(self):

        nom = self.champ_nom.get().strip()

        annee = self.champ_annee.get().strip()

        if not nom:

            messagebox.showwarning(
                "Information manquante",
                "Veuillez saisir le nom de l'établissement."
            )

            return

        if not annee:

            messagebox.showwarning(
                "Information manquante",
                "Veuillez saisir l'année scolaire."
            )

            return

        try:

            enregistrer_ecole(
                nom,
                self.type_var.get(),
                self.champ_adresse.get().strip(),
                self.champ_localite.get().strip(),
                self.champ_telephone.get().strip(),
                self.champ_email.get().strip(),
                self.champ_dre.get().strip(),
                self.champ_directeur.get().strip(),
                self.champ_secretaire.get().strip(),
                self.champ_devise.get().strip(),
                self.logo_selectionne,
                annee
            )

            self.charger_donnees()

            self.logo_selectionne = ""

            messagebox.showinfo(
                "Configuration terminée",
                "Les informations de l'établissement "
                "ont été enregistrées avec succès."
            )

            self.construire_interface_principale()

        except Exception as erreur:

            messagebox.showerror(
                "Erreur",
                f"Impossible d'enregistrer la configuration.\n\n{erreur}"
            )


    # ========================================================
    # INTERFACE PRINCIPALE
    # ========================================================

    def construire_interface_principale(self):

        for widget in self.winfo_children():
            widget.destroy()

        self.menu_buttons = {}

        self.creer_interface()

        self.afficher_accueil()


    def creer_interface(self):

        self.title("GestionScolaire")

        self.menu = tk.Frame(
            self,
            bg=MENU,
            width=235
        )

        self.menu.pack(
            side="left",
            fill="y"
        )

        self.menu.pack_propagate(False)

        # ----------------------------------------------------
        # IDENTITÉ
        # ----------------------------------------------------

        identite = tk.Frame(
            self.menu,
            bg=MENU,
            height=120
        )

        identite.pack(fill="x")

        identite.pack_propagate(False)

        tk.Label(
            identite,
            text="GestionScolaire",
            bg=MENU,
            fg=BLANC,
            font=("Segoe UI", 18, "bold")
        ).pack(
            anchor="w",
            padx=20,
            pady=(14, 0)
        )

        tk.Label(
            identite,
            text=self.etablissement["nom"],
            bg=MENU,
            fg=TURQUOISE_VIF,
            font=("Segoe UI", 12, "bold")
        ).pack(
            anchor="w",
            padx=21,
            pady=(7, 0)
        )

        tk.Label(
            identite,
            text=(
                self.etablissement["type_etablissement"]
                or ""
            ),
            bg=MENU,
            fg="#CBD5E1",
            font=("Segoe UI", 8)
        ).pack(
            anchor="w",
            padx=21,
            pady=(2, 0)
        )

        tk.Frame(
            self.menu,
            bg="#3A4A5A",
            height=1
        ).pack(
            fill="x",
            padx=15,
            pady=8
        )

        # ----------------------------------------------------
        # MENU
        # ----------------------------------------------------

        self.creer_bouton_menu(
            "⌂",
            "ACCUEIL",
            self.afficher_accueil
        )

        self.creer_bouton_menu(
            "▣",
            "L'ÉCOLE",
            self.afficher_ecole
        )

        self.creer_bouton_menu(
            "◷",
            "EMPLOI DU TEMPS",
            self.afficher_emploi
        )

        self.creer_bouton_menu(
            "▤",
            "CONSEIL",
            self.afficher_conseil
        )

        self.creer_bouton_menu(
            "⚙",
            "DÉVELOPPEUR",
            self.afficher_developpeur
        )

        # ----------------------------------------------------
        # ZONE PRINCIPALE
        # ----------------------------------------------------

        self.zone_principale = tk.Frame(
            self,
            bg=FOND
        )

        self.zone_principale.pack(
            side="left",
            fill="both",
            expand=True
        )

        # ----------------------------------------------------
        # BARRE SUPÉRIEURE
        # ----------------------------------------------------

        self.barre = tk.Frame(
            self.zone_principale,
            bg=TURQUOISE,
            height=68
        )

        self.barre.pack(fill="x")

        self.barre.pack_propagate(False)

        self.titre = tk.Label(
            self.barre,
            text="TABLEAU DE BORD",
            bg=TURQUOISE,
            fg=BLANC,
            font=("Segoe UI", 16, "bold")
        )

        self.titre.pack(
            side="left",
            padx=25
        )

        annee = (
            self.annee_active["libelle"]
            if self.annee_active
            else "Non définie"
        )

        tk.Label(
            self.barre,
            text=annee,
            bg=TURQUOISE,
            fg=BLANC,
            font=("Segoe UI", 10, "bold")
        ).pack(
            side="right",
            padx=25
        )

        tk.Label(
            self.barre,
            text="Année scolaire",
            bg=TURQUOISE,
            fg="#D5F5F2",
            font=("Segoe UI", 9)
        ).pack(
            side="right"
        )

        self.contenu = tk.Frame(
            self.zone_principale,
            bg=FOND
        )

        self.contenu.pack(
            fill="both",
            expand=True
        )


    # ========================================================
    # BOUTON MENU
    # ========================================================

    def creer_bouton_menu(
        self,
        icone,
        texte,
        commande
    ):

        bouton = tk.Frame(
            self.menu,
            bg=MENU,
            height=48,
            cursor="hand2"
        )

        bouton.pack(
            fill="x",
            padx=10,
            pady=2
        )

        bouton.pack_propagate(False)

        icone_label = tk.Label(
            bouton,
            text=icone,
            bg=MENU,
            fg="#CBD5E1",
            font=("Segoe UI Symbol", 15),
            width=3
        )

        icone_label.pack(
            side="left",
            padx=(4, 0)
        )

        texte_label = tk.Label(
            bouton,
            text=texte,
            bg=MENU,
            fg=BLANC,
            font=("Segoe UI", 10, "bold")
        )

        texte_label.pack(
            side="left"
        )

        self.menu_buttons[texte] = (
            bouton,
            icone_label,
            texte_label
        )

        # ----------------------------------------------------
        # CLIC
        # ----------------------------------------------------

        for widget in (
            bouton,
            icone_label,
            texte_label
        ):

            widget.bind(
                "<Button-1>",
                lambda event, f=commande: f()
            )

        # ----------------------------------------------------
        # SURVOL
        # ----------------------------------------------------

        for widget in (
            bouton,
            icone_label,
            texte_label
        ):

            widget.bind(
                "<Enter>",
                lambda event,
                b=bouton,
                i=icone_label,
                t=texte_label:
                self.menu_survol(b, i, t)
            )

            widget.bind(
                "<Leave>",
                lambda event,
                b=bouton,
                i=icone_label,
                t=texte_label:
                self.menu_quitte(b, i, t)
            )


    def menu_survol(
        self,
        bouton,
        icone,
        texte
    ):

        if bouton["bg"] != MENU_ACTIF:

            bouton.configure(
                bg=MENU_HOVER
            )

            icone.configure(
                bg=MENU_HOVER
            )

            texte.configure(
                bg=MENU_HOVER
            )


    def menu_quitte(
        self,
        bouton,
        icone,
        texte
    ):

        if bouton["bg"] != MENU_ACTIF:

            bouton.configure(
                bg=MENU
            )

            icone.configure(
                bg=MENU
            )

            texte.configure(
                bg=MENU
            )


    def activer_menu(self, nom):

        for nom_menu, elements in self.menu_buttons.items():

            bouton, icone, texte = elements

            if nom_menu == nom:

                bouton.configure(
                    bg=MENU_ACTIF
                )

                icone.configure(
                    bg=MENU_ACTIF
                )

                texte.configure(
                    bg=MENU_ACTIF
                )

            else:

                bouton.configure(
                    bg=MENU
                )

                icone.configure(
                    bg=MENU
                )

                texte.configure(
                    bg=MENU
                )


    # ========================================================
    # NETTOYAGE
    # ========================================================

    def nettoyer(self):

        for widget in self.contenu.winfo_children():
            widget.destroy()


    # ========================================================
    # ACCUEIL
    # ========================================================

    def afficher_accueil(self):

        self.nettoyer()

        self.activer_menu("ACCUEIL")

        self.titre.configure(
            text="TABLEAU DE BORD"
        )

        cadre = tk.Frame(
            self.contenu,
            bg=FOND
        )

        cadre.pack(
            fill="both",
            expand=True,
            padx=25,
            pady=22
        )

        tk.Label(
            cadre,
            text="Tableau de bord",
            bg=FOND,
            fg=TEXTE,
            font=("Segoe UI", 22, "bold")
        ).pack(anchor="w")

        tk.Label(
            cadre,
            text=self.etablissement["nom"],
            bg=FOND,
            fg=TURQUOISE,
            font=("Segoe UI", 11, "bold")
        ).pack(
            anchor="w",
            pady=(3, 18)
        )

        cartes = tk.Frame(
            cadre,
            bg=FOND
        )

        cartes.pack(
            fill="x"
        )

        self.creer_carte(
            cartes,
            "ÉLÈVES",
            "0",
            BLEU,
            "Effectif total"
        )

        self.creer_carte(
            cartes,
            "FILLES",
            "0",
            VERT,
            "Effectif féminin"
        )

        self.creer_carte(
            cartes,
            "GARÇONS",
            "0",
            BLEU,
            "Effectif masculin"
        )

        self.creer_carte(
            cartes,
            "CLASSES",
            "0",
            ORANGE,
            "Classes ouvertes"
        )

        bas = tk.Frame(
            cadre,
            bg=FOND
        )

        bas.pack(
            fill="both",
            expand=True,
            pady=(20, 0)
        )

        # ----------------------------------------------------
        # INFORMATIONS
        # ----------------------------------------------------

        bloc_infos = tk.Frame(
            bas,
            bg=INFORMATIONS_FOND,
            highlightbackground=BORDURE,
            highlightthickness=1
        )

        bloc_infos.pack(
            side="left",
            fill="both",
            expand=True,
            padx=5
        )

        tk.Frame(
            bloc_infos,
            bg=BLEU,
            height=5
        ).pack(fill="x")

        tk.Label(
            bloc_infos,
            text="INFORMATIONS SCOLAIRES",
            bg=INFORMATIONS_FOND,
            fg=TEXTE,
            font=("Segoe UI", 12, "bold")
        ).pack(
            anchor="w",
            padx=18,
            pady=(14, 8)
        )

        nom = (
            self.etablissement["nom"]
            or "Non renseigné"
        )

        type_ecole = (
            self.etablissement["type_etablissement"]
            or "Non renseigné"
        )

        directeur = (
            self.etablissement["directeur"]
            or "Non renseigné"
        )

        secretaire = (
            self.etablissement["secretaire"]
            or "Non renseigné"
        )

        localite = (
            self.etablissement["localite"]
            or "Non renseignée"
        )

        annee = (
            self.annee_active["libelle"]
            if self.annee_active
            else "Non définie"
        )

        details = tk.Frame(
            bloc_infos,
            bg=INFORMATIONS_FOND
        )

        details.pack(
            fill="x",
            padx=18
        )

        tk.Label(
            details,
            text=nom,
            bg=INFORMATIONS_FOND,
            fg=TURQUOISE,
            font=("Segoe UI", 14, "bold")
        ).pack(anchor="w")

        tk.Label(
            details,
            text=f"Type : {type_ecole}",
            bg=INFORMATIONS_FOND,
            fg=TEXTE_SECONDAIRE,
            font=("Segoe UI", 9)
        ).pack(
            anchor="w",
            pady=(4, 0)
        )

        tk.Label(
            details,
            text=f"Localité : {localite}",
            bg=INFORMATIONS_FOND,
            fg=TEXTE_SECONDAIRE,
            font=("Segoe UI", 9)
        ).pack(
            anchor="w",
            pady=(3, 0)
        )

        tk.Label(
            details,
            text=f"Directeur : {directeur}",
            bg=INFORMATIONS_FOND,
            fg=TEXTE_SECONDAIRE,
            font=("Segoe UI", 9)
        ).pack(
            anchor="w",
            pady=(3, 0)
        )

        tk.Label(
            details,
            text=f"Secrétaire : {secretaire}",
            bg=INFORMATIONS_FOND,
            fg=TEXTE_SECONDAIRE,
            font=("Segoe UI", 9)
        ).pack(
            anchor="w",
            pady=(3, 0)
        )

        tk.Label(
            details,
            text=f"Année scolaire : {annee}",
            bg=INFORMATIONS_FOND,
            fg=TEXTE_SECONDAIRE,
            font=("Segoe UI", 9)
        ).pack(
            anchor="w",
            pady=(3, 0)
        )

        # ----------------------------------------------------
        # ALERTES
        # ----------------------------------------------------

        bloc_alertes = tk.Frame(
            bas,
            bg=ALERTES_FOND,
            highlightbackground=BORDURE,
            highlightthickness=1
        )

        bloc_alertes.pack(
            side="left",
            fill="both",
            expand=True,
            padx=5
        )

        tk.Frame(
            bloc_alertes,
            bg=ORANGE,
            height=5
        ).pack(fill="x")

        tk.Label(
            bloc_alertes,
            text="ALERTES",
            bg=ALERTES_FOND,
            fg=TEXTE,
            font=("Segoe UI", 12, "bold")
        ).pack(
            anchor="w",
            padx=18,
            pady=(14, 8)
        )

        tk.Label(
            bloc_alertes,
            text="Aucune alerte pour le moment.",
            bg=ALERTES_FOND,
            fg=TEXTE_SECONDAIRE,
            font=("Segoe UI", 10)
        ).pack(
            anchor="w",
            padx=18
        )


    # ========================================================
    # L'ÉCOLE
    # ========================================================

    def afficher_ecole(self):

        self.nettoyer()

        self.activer_menu("L'ÉCOLE")

        self.titre.configure(
            text="L'ÉCOLE"
        )

        cadre = tk.Frame(
            self.contenu,
            bg=FOND
        )

        cadre.pack(
            fill="both",
            expand=True,
            padx=25,
            pady=22
        )

        tk.Label(
            cadre,
            text="L'ÉCOLE",
            bg=FOND,
            fg=TEXTE,
            font=("Segoe UI", 22, "bold")
        ).pack(anchor="w")

        tk.Label(
            cadre,
            text="Gestion de l'établissement",
            bg=FOND,
            fg=TEXTE_SECONDAIRE,
            font=("Segoe UI", 10)
        ).pack(
            anchor="w",
            pady=(3, 20)
        )

        grille = tk.Frame(
            cadre,
            bg=FOND
        )

        grille.pack(
            fill="both",
            expand=True
        )

        # ----------------------------------------------------
        # ÉLÈVES
        # ----------------------------------------------------

        self.creer_module(
            grille,
            "ÉLÈVES",
            "Gestion des élèves",
            BLEU,
            0,
            0,
            self.module_eleves
        )

        # ----------------------------------------------------
        # ENSEIGNANTS
        # ----------------------------------------------------

        self.creer_module(
            grille,
            "ENSEIGNANTS",
            "Dossiers et affectations",
            VERT,
            0,
            1,
            self.module_enseignants
        )

        # ----------------------------------------------------
        # CLASSES
        # ----------------------------------------------------

        self.creer_module(
            grille,
            "CLASSES",
            "Classes et niveaux",
            ORANGE,
            1,
            0,
            self.module_classes
        )

        # ----------------------------------------------------
        # MATIÈRES
        # ----------------------------------------------------

        self.creer_module(
            grille,
            "MATIÈRES",
            "Matières et coefficients",
            BLEU,
            1,
            1,
            self.module_matieres
        )


    # ========================================================
    # MODULE ÉLÈVES
    # ========================================================

    def module_eleves(self):
        self.nettoyer()
        self.activer_menu("L'ÉCOLE")
        self.titre.configure(text="ÉLÈVES")

        cadre = tk.Frame(self.contenu, bg=FOND)
        cadre.pack(fill="both", expand=True, padx=25, pady=18)

        entete = tk.Frame(cadre, bg=FOND)
        entete.pack(fill="x")

        bloc_titre = tk.Frame(entete, bg=FOND)
        bloc_titre.pack(side="left", fill="x", expand=True)

        tk.Label(
            bloc_titre, text="Gestion des élèves", bg=FOND, fg=TEXTE,
            font=("Segoe UI", 20, "bold")
        ).pack(anchor="w")

        annee = self.annee_active["libelle"] if self.annee_active else "Aucune année active"
        tk.Label(
            bloc_titre,
            text=f"Dossiers administratifs • Année scolaire : {annee}",
            bg=FOND, fg=TEXTE_SECONDAIRE, font=("Segoe UI", 9)
        ).pack(anchor="w", pady=(3, 0))

        tk.Button(
            entete, text="+ NOUVEL ÉLÈVE", command=self.ouvrir_formulaire_eleve,
            bg=TURQUOISE, fg=BLANC, activebackground="#115E59",
            relief="flat", bd=0, cursor="hand2",
            font=("Segoe UI", 10, "bold"), padx=18, pady=10
        ).pack(side="right")

        stats = self.obtenir_stats_eleves()
        cartes = tk.Frame(cadre, bg=FOND)
        cartes.pack(fill="x", pady=(18, 12))
        self.creer_carte(cartes, "ÉLÈVES", str(stats["total"]), BLEU, "Dossiers enregistrés")
        self.creer_carte(cartes, "FILLES", str(stats["filles"]), VERT, "Effectif féminin")
        self.creer_carte(cartes, "GARÇONS", str(stats["garcons"]), BLEU, "Effectif masculin")
        self.creer_carte(cartes, "ACTIFS", str(stats["actifs"]), ORANGE, "Situation actuelle")

        barre = tk.Frame(cadre, bg=CARTE, highlightbackground=BORDURE, highlightthickness=1)
        barre.pack(fill="x", pady=(0, 10))

        tk.Label(
            barre, text="Rechercher :", bg=CARTE, fg=TEXTE,
            font=("Segoe UI", 9, "bold")
        ).pack(side="left", padx=(14, 8), pady=10)

        self.recherche_eleve = tk.StringVar()
        recherche = tk.Entry(
            barre, textvariable=self.recherche_eleve, bg=BLANC, fg=TEXTE,
            insertbackground=TEXTE, relief="solid", bd=1, font=("Segoe UI", 10)
        )
        recherche.pack(side="left", fill="x", expand=True, padx=(0, 8), ipady=6)
        recherche.bind("<KeyRelease>", lambda event: self.charger_liste_eleves())

        tk.Button(
            barre, text="EFFACER",
            command=lambda: self.effacer_recherche_eleve(),
            bg="#E2E8F0", fg=TEXTE, activebackground="#CBD5E1",
            relief="flat", bd=0, cursor="hand2",
            font=("Segoe UI", 9, "bold"), padx=14, pady=8
        ).pack(side="right", padx=(0, 8))

        zone = tk.Frame(cadre, bg=CARTE, highlightbackground=BORDURE, highlightthickness=1)
        zone.pack(fill="both", expand=True)

        colonnes = ("matricule", "nom", "prenom", "sexe", "date", "statut", "situation")
        self.table_eleves = ttk.Treeview(zone, columns=colonnes, show="headings", selectmode="browse")

        titres = {
            "matricule": "MATRICULE", "nom": "NOM", "prenom": "PRÉNOM(S)",
            "sexe": "SEXE", "date": "DATE DE NAISSANCE",
            "statut": "STATUT D'ENTRÉE", "situation": "SITUATION ACTUELLE"
        }
        largeurs = {
            "matricule": 120, "nom": 150, "prenom": 170, "sexe": 75,
            "date": 135, "statut": 145, "situation": 145
        }

        for colonne in colonnes:
            self.table_eleves.heading(colonne, text=titres[colonne])
            self.table_eleves.column(colonne, width=largeurs[colonne], minwidth=70, anchor="w")

        sy = ttk.Scrollbar(zone, orient="vertical", command=self.table_eleves.yview)
        sx = ttk.Scrollbar(zone, orient="horizontal", command=self.table_eleves.xview)
        self.table_eleves.configure(yscrollcommand=sy.set, xscrollcommand=sx.set)
        sy.pack(side="right", fill="y")
        sx.pack(side="bottom", fill="x")
        self.table_eleves.pack(side="left", fill="both", expand=True)
        self.table_eleves.bind("<Double-1>", lambda event: self.modifier_eleve())

        pied = tk.Frame(cadre, bg=FOND)
        pied.pack(fill="x", pady=(10, 0))

        self.eleve_selection_info = tk.Label(
            pied, text="", bg=FOND, fg=TEXTE_SECONDAIRE, font=("Segoe UI", 9)
        )
        self.eleve_selection_info.pack(side="left")

        tk.Button(
            pied, text="SUPPRIMER", command=self.supprimer_eleve,
            bg=ROUGE, fg=BLANC, activebackground="#B91C1C",
            relief="flat", bd=0, cursor="hand2",
            font=("Segoe UI", 9, "bold"), padx=15, pady=8
        ).pack(side="right", padx=(8, 0))

        tk.Button(
            pied, text="MODIFIER", command=self.modifier_eleve,
            bg=BLEU, fg=BLANC, activebackground="#2563EB",
            relief="flat", bd=0, cursor="hand2",
            font=("Segoe UI", 9, "bold"), padx=15, pady=8
        ).pack(side="right")

        self.charger_liste_eleves()

    def effacer_recherche_eleve(self):
        self.recherche_eleve.set("")
        self.charger_liste_eleves()

    def obtenir_stats_eleves(self):
        db = connexion()
        c = db.cursor()
        c.execute("SELECT COUNT(*) AS n FROM eleves")
        total = c.fetchone()["n"]
        c.execute("SELECT COUNT(*) AS n FROM eleves WHERE sexe = 'Féminin'")
        filles = c.fetchone()["n"]
        c.execute("SELECT COUNT(*) AS n FROM eleves WHERE sexe = 'Masculin'")
        garcons = c.fetchone()["n"]
        c.execute("SELECT COUNT(*) AS n FROM eleves WHERE situation_actuelle = 'Actif'")
        actifs = c.fetchone()["n"]
        db.close()
        return {"total": total, "filles": filles, "garcons": garcons, "actifs": actifs}

    def charger_liste_eleves(self):
        if not hasattr(self, "table_eleves"):
            return

        for item in self.table_eleves.get_children():
            self.table_eleves.delete(item)

        recherche = self.recherche_eleve.get().strip()
        db = connexion()
        c = db.cursor()

        if recherche:
            terme = f"%{recherche}%"
            c.execute("""
                SELECT id, matricule, nom, prenom, sexe, date_naissance,
                       statut_entree, situation_actuelle
                FROM eleves
                WHERE matricule LIKE ? OR nom LIKE ? OR prenom LIKE ?
                ORDER BY nom COLLATE NOCASE, prenom COLLATE NOCASE
            """, (terme, terme, terme))
        else:
            c.execute("""
                SELECT id, matricule, nom, prenom, sexe, date_naissance,
                       statut_entree, situation_actuelle
                FROM eleves
                ORDER BY nom COLLATE NOCASE, prenom COLLATE NOCASE
            """)

        lignes = c.fetchall()
        db.close()

        for ligne in lignes:
            self.table_eleves.insert(
                "", "end", iid=str(ligne["id"]),
                values=(
                    ligne["matricule"], ligne["nom"], ligne["prenom"],
                    ligne["sexe"], ligne["date_naissance"] or "",
                    ligne["statut_entree"], ligne["situation_actuelle"]
                )
            )

        self.eleve_selection_info.configure(
            text=f"{len(lignes)} élève(s) affiché(s). Double-cliquez sur une ligne pour modifier."
        )

    def ouvrir_formulaire_eleve(self, eleve_id=None):
        fenetre = tk.Toplevel(self)
        fenetre.title("Nouvel élève" if eleve_id is None else "Modifier l'élève")
        fenetre.geometry("720x650")
        fenetre.minsize(680, 600)
        fenetre.configure(bg=FOND)
        fenetre.transient(self)
        fenetre.grab_set()

        entete = tk.Frame(fenetre, bg=TURQUOISE, height=75)
        entete.pack(fill="x")
        entete.pack_propagate(False)

        tk.Label(
            entete,
            text="NOUVEL ÉLÈVE" if eleve_id is None else "MODIFICATION DE L'ÉLÈVE",
            bg=TURQUOISE, fg=BLANC, font=("Segoe UI", 17, "bold")
        ).pack(anchor="w", padx=25, pady=(15, 0))

        tk.Label(
            entete, text="Dossier administratif et situation scolaire",
            bg=TURQUOISE, fg="#D5F5F2", font=("Segoe UI", 9)
        ).pack(anchor="w", padx=25)

        zone = tk.Frame(fenetre, bg=FOND)
        zone.pack(fill="both", expand=True, padx=25, pady=18)

        carte = tk.Frame(zone, bg=CARTE, highlightbackground=BORDURE, highlightthickness=1)
        carte.pack(fill="both", expand=True)

        interieur = tk.Frame(carte, bg=CARTE)
        interieur.pack(fill="both", expand=True, padx=22, pady=18)

        champs = {}

        def champ(parent, nom, titre):
            bloc = tk.Frame(parent, bg=CARTE)
            bloc.pack(side="left", fill="x", expand=True, padx=5)
            tk.Label(
                bloc, text=titre, bg=CARTE, fg=TEXTE,
                font=("Segoe UI", 9, "bold")
            ).pack(anchor="w", pady=(0, 4))
            e = tk.Entry(
                bloc, bg=BLANC, fg=TEXTE, insertbackground=TEXTE,
                relief="solid", bd=1, font=("Segoe UI", 10)
            )
            e.pack(fill="x", ipady=6)
            champs[nom] = e
            return e

        def combo(parent, titre, valeurs, valeur):
            bloc = tk.Frame(parent, bg=CARTE)
            bloc.pack(side="left", fill="x", expand=True, padx=5)
            tk.Label(
                bloc, text=titre, bg=CARTE, fg=TEXTE,
                font=("Segoe UI", 9, "bold")
            ).pack(anchor="w", pady=(0, 4))
            var = tk.StringVar(value=valeur)
            w = ttk.Combobox(
                bloc, textvariable=var, values=valeurs,
                state="readonly", font=("Segoe UI", 10)
            )
            w.pack(fill="x", ipady=5)
            return var

        ligne = tk.Frame(interieur, bg=CARTE)
        ligne.pack(fill="x", pady=(0, 10))
        champ(ligne, "matricule", "Matricule *")
        champ(ligne, "nom", "Nom *")

        ligne = tk.Frame(interieur, bg=CARTE)
        ligne.pack(fill="x", pady=(0, 10))
        champ(ligne, "prenom", "Prénom(s) *")
        champ(ligne, "date_naissance", "Date de naissance")

        ligne = tk.Frame(interieur, bg=CARTE)
        ligne.pack(fill="x", pady=(0, 10))
        champ(ligne, "lieu_naissance", "Lieu de naissance")
        champ(ligne, "nationalite", "Nationalité")

        ligne = tk.Frame(interieur, bg=CARTE)
        ligne.pack(fill="x", pady=(0, 10))
        sexe_var = combo(ligne, "Sexe *", ("Féminin", "Masculin"), "Féminin")
        statut_var = combo(
            ligne, "Statut d'entrée *",
            ("Nouveau", "Ancien", "Transféré", "Réinscrit"), "Nouveau"
        )

        ligne = tk.Frame(interieur, bg=CARTE)
        ligne.pack(fill="x", pady=(0, 10))
        situation_var = combo(
            ligne, "Situation actuelle *",
            ("Actif", "Abandonné", "Exclu", "Transféré", "Sorti", "Décédé"),
            "Actif"
        )
        tk.Frame(ligne, bg=CARTE).pack(side="left", fill="x", expand=True, padx=5)

        ligne = tk.Frame(interieur, bg=CARTE)
        ligne.pack(fill="x", pady=(0, 10))
        champ(ligne, "nom_parent", "Parent / Responsable")
        champ(ligne, "telephone_parent", "Téléphone du responsable")

        tk.Label(
            interieur, text="Observation", bg=CARTE, fg=TEXTE,
            font=("Segoe UI", 9, "bold")
        ).pack(anchor="w", padx=5, pady=(3, 4))

        observation = tk.Text(
            interieur, height=5, bg=BLANC, fg=TEXTE,
            insertbackground=TEXTE, relief="solid", bd=1,
            font=("Segoe UI", 10)
        )
        observation.pack(fill="x", padx=5)

        if eleve_id is not None:
            db = connexion()
            c = db.cursor()
            c.execute("SELECT * FROM eleves WHERE id = ?", (eleve_id,))
            eleve = c.fetchone()
            db.close()

            if not eleve:
                fenetre.destroy()
                messagebox.showerror("Erreur", "L'élève demandé n'existe plus.", parent=self)
                return

            for nom in champs:
                champs[nom].insert(0, eleve[nom] or "")
            sexe_var.set(eleve["sexe"] or "Féminin")
            statut_var.set(eleve["statut_entree"] or "Nouveau")
            situation_var.set(eleve["situation_actuelle"] or "Actif")
            observation.insert("1.0", eleve["observation"] or "")

        boutons = tk.Frame(zone, bg=FOND)
        boutons.pack(fill="x", pady=(12, 0))

        tk.Button(
            boutons, text="ANNULER", command=fenetre.destroy,
            bg="#E2E8F0", fg=TEXTE, activebackground="#CBD5E1",
            relief="flat", bd=0, cursor="hand2",
            font=("Segoe UI", 9, "bold"), padx=18, pady=9
        ).pack(side="left")

        def enregistrer():
            matricule = champs["matricule"].get().strip()
            nom = champs["nom"].get().strip()
            prenom = champs["prenom"].get().strip()

            if not matricule or not nom or not prenom:
                messagebox.showwarning(
                    "Informations manquantes",
                    "Le matricule, le nom et le prénom sont obligatoires.",
                    parent=fenetre
                )
                return

            valeurs = (
                matricule, nom, prenom, sexe_var.get(),
                champs["date_naissance"].get().strip(),
                champs["lieu_naissance"].get().strip(),
                champs["nationalite"].get().strip(),
                statut_var.get(), situation_var.get(),
                champs["nom_parent"].get().strip(),
                champs["telephone_parent"].get().strip(),
                observation.get("1.0", "end").strip()
            )

            db = None
            try:
                db = connexion()
                c = db.cursor()

                if eleve_id is None:
                    c.execute("""
                        INSERT INTO eleves (
                            matricule, nom, prenom, sexe, date_naissance,
                            lieu_naissance, nationalite, statut_entree,
                            situation_actuelle, nom_parent, telephone_parent,
                            observation
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """, valeurs)

                    nouvel_id = c.lastrowid
                    if self.annee_active:
                        c.execute("""
                            INSERT OR IGNORE INTO inscriptions_eleves
                            (eleve_id, annee_id, statut)
                            VALUES (?, ?, 'Inscrit')
                        """, (nouvel_id, self.annee_active["id"]))
                else:
                    c.execute("""
                        UPDATE eleves SET
                            matricule=?, nom=?, prenom=?, sexe=?,
                            date_naissance=?, lieu_naissance=?, nationalite=?,
                            statut_entree=?, situation_actuelle=?,
                            nom_parent=?, telephone_parent=?, observation=?
                        WHERE id=?
                    """, valeurs + (eleve_id,))

                db.commit()
                db.close()
                fenetre.destroy()
                self.module_eleves()
                messagebox.showinfo(
                    "Enregistrement réussi",
                    "Le dossier de l'élève a été enregistré avec succès.",
                    parent=self
                )

            except sqlite3.IntegrityError:
                if db:
                    db.close()
                messagebox.showerror(
                    "Matricule déjà utilisé",
                    "Ce matricule existe déjà. Veuillez en choisir un autre.",
                    parent=fenetre
                )
            except Exception as erreur:
                if db:
                    db.close()
                messagebox.showerror(
                    "Erreur",
                    f"Impossible d'enregistrer l'élève.\n\n{erreur}",
                    parent=fenetre
                )

        tk.Button(
            boutons, text="ENREGISTRER", command=enregistrer,
            bg=TURQUOISE, fg=BLANC, activebackground="#115E59",
            relief="flat", bd=0, cursor="hand2",
            font=("Segoe UI", 10, "bold"), padx=22, pady=10
        ).pack(side="right")

        champs["matricule"].focus_set()

    def modifier_eleve(self):
        selection = self.table_eleves.selection()
        if not selection:
            messagebox.showwarning(
                "Aucune sélection",
                "Veuillez sélectionner un élève à modifier."
            )
            return
        self.ouvrir_formulaire_eleve(int(selection[0]))

    def supprimer_eleve(self):
        selection = self.table_eleves.selection()
        if not selection:
            messagebox.showwarning(
                "Aucune sélection",
                "Veuillez sélectionner un élève à supprimer."
            )
            return

        eleve_id = int(selection[0])
        valeurs = self.table_eleves.item(selection[0], "values")
        identite = f"{valeurs[0]} — {valeurs[1]} {valeurs[2]}"

        if not messagebox.askyesno(
            "Confirmer la suppression",
            f"Voulez-vous supprimer le dossier de :\n\n{identite} ?\n\n"
            "Ses inscriptions annuelles seront également supprimées.",
            icon="warning"
        ):
            return

        db = None
        try:
            db = connexion()
            c = db.cursor()
            c.execute("DELETE FROM inscriptions_eleves WHERE eleve_id=?", (eleve_id,))
            c.execute("DELETE FROM eleves WHERE id=?", (eleve_id,))
            db.commit()
            db.close()
            self.module_eleves()
            messagebox.showinfo("Suppression réussie", "Le dossier de l'élève a été supprimé.")
        except Exception as erreur:
            if db:
                db.close()
            messagebox.showerror(
                "Erreur",
                f"Impossible de supprimer l'élève.\n\n{erreur}"
            )


    # ========================================================
    # MODULE ENSEIGNANTS
    # ========================================================

    def module_enseignants(self):

        self.afficher_page(
            "ENSEIGNANTS",
            "Dossiers administratifs et affectations pédagogiques.",
            "Le module de gestion des enseignants sera construit ici.",
            "L'ÉCOLE"
        )


    # ========================================================
    # MODULE CLASSES
    # ========================================================

    def module_classes(self):
        self.nettoyer()
        self.activer_menu("L'ÉCOLE")
        self.titre.configure(text="CLASSES")

        cadre = tk.Frame(self.contenu, bg=FOND)
        cadre.pack(fill="both", expand=True, padx=25, pady=18)

        entete = tk.Frame(cadre, bg=FOND)
        entete.pack(fill="x")
        bloc = tk.Frame(entete, bg=FOND)
        bloc.pack(side="left", fill="x", expand=True)
        tk.Label(bloc, text="Gestion des classes", bg=FOND, fg=TEXTE,
                 font=("Segoe UI", 20, "bold")).pack(anchor="w")
        annee = self.annee_active["libelle"] if self.annee_active else "Aucune année active"
        tk.Label(bloc, text=f"Structure des classes • Année scolaire : {annee}",
                 bg=FOND, fg=TEXTE_SECONDAIRE, font=("Segoe UI", 9)).pack(anchor="w", pady=(3, 0))

        tk.Button(entete, text="+ NOUVELLE CLASSE", command=self.ouvrir_formulaire_classe,
                  bg=TURQUOISE, fg=BLANC, activebackground="#115E59", relief="flat", bd=0,
                  cursor="hand2", font=("Segoe UI", 10, "bold"), padx=18, pady=10).pack(side="right")

        barre = tk.Frame(cadre, bg=CARTE, highlightbackground=BORDURE, highlightthickness=1)
        barre.pack(fill="x", pady=(18, 10))
        tk.Label(barre, text="Rechercher :", bg=CARTE, fg=TEXTE,
                 font=("Segoe UI", 9, "bold")).pack(side="left", padx=(14, 8), pady=10)
        self.recherche_classe = tk.StringVar()
        recherche = tk.Entry(barre, textvariable=self.recherche_classe, bg=BLANC, fg=TEXTE,
                             insertbackground=TEXTE, relief="solid", bd=1, font=("Segoe UI", 10))
        recherche.pack(side="left", fill="x", expand=True, padx=(0, 8), ipady=6)
        recherche.bind("<KeyRelease>", lambda e: self.charger_liste_classes())
        tk.Button(barre, text="EFFACER", command=self.effacer_recherche_classe,
                  bg="#E2E8F0", fg=TEXTE, relief="flat", bd=0, cursor="hand2",
                  font=("Segoe UI", 9, "bold"), padx=14, pady=8).pack(side="right", padx=8)
        tk.Button(barre, text="MODÈLE EXCEL", command=self.creer_modele_excel_classe,
                  bg="#64748B", fg=BLANC, activebackground="#475569", relief="flat", bd=0,
                  cursor="hand2", font=("Segoe UI", 9, "bold"), padx=14, pady=8).pack(side="right", padx=(0, 8))
        tk.Button(barre, text="IMPORTER EXCEL", command=self.importer_eleves_classe_excel,
                  bg=VERT, fg=BLANC, activebackground="#18794E", relief="flat", bd=0,
                  cursor="hand2", font=("Segoe UI", 9, "bold"), padx=14, pady=8).pack(side="right", padx=(0, 8))

        zone = tk.Frame(cadre, bg=CARTE, highlightbackground=BORDURE, highlightthickness=1)
        zone.pack(fill="both", expand=True)
        cols = ("code", "niveau", "libelle", "effectif", "maximum", "statut")
        self.table_classes = ttk.Treeview(zone, columns=cols, show="headings", selectmode="browse")
        titres = {"code":"CODE", "niveau":"NIVEAU", "libelle":"CLASSE", "effectif":"EFFECTIF", "maximum":"MAXIMUM", "statut":"STATUT"}
        widths = {"code":120, "niveau":140, "libelle":180, "effectif":100, "maximum":100, "statut":100}
        for col in cols:
            self.table_classes.heading(col, text=titres[col])
            self.table_classes.column(col, width=widths[col], minwidth=75,
                                      anchor="center" if col in ("effectif", "maximum", "statut") else "w")
        sy = ttk.Scrollbar(zone, orient="vertical", command=self.table_classes.yview)
        sy.pack(side="right", fill="y")
        self.table_classes.configure(yscrollcommand=sy.set)
        self.table_classes.pack(side="left", fill="both", expand=True)
        self.table_classes.bind("<Double-1>", lambda e: self.afficher_eleves_classe_selectionnee())

        pied = tk.Frame(cadre, bg=FOND)
        pied.pack(fill="x", pady=(10, 0))
        self.classe_selection_info = tk.Label(pied, text="", bg=FOND, fg=TEXTE_SECONDAIRE, font=("Segoe UI", 9))
        self.classe_selection_info.pack(side="left")
        for txt, cmd, bg in [
            ("ÉLÈVES DE LA CLASSE", self.afficher_eleves_classe_selectionnee, BLEU),
            ("EXPORTER EXCEL", self.exporter_eleves_classe_excel, VERT),
            ("MODIFIER", self.modifier_classe, BLEU),
            ("SUPPRIMER", self.supprimer_classe, ROUGE)
        ]:
            tk.Button(pied, text=txt, command=cmd, bg=bg, fg=BLANC,
                      activebackground="#2563EB" if bg == BLEU else ("#18794E" if bg == VERT else "#B91C1C"),
                      relief="flat", bd=0, cursor="hand2", font=("Segoe UI", 9, "bold"),
                      padx=15, pady=8).pack(side="right", padx=(8, 0))
        self.charger_liste_classes()

    def effacer_recherche_classe(self):
        self.recherche_classe.set("")
        self.charger_liste_classes()

    def charger_liste_classes(self):
        if not hasattr(self, "table_classes"):
            return
        for item in self.table_classes.get_children():
            self.table_classes.delete(item)
        recherche = self.recherche_classe.get().strip()
        db = connexion()
        c = db.cursor()
        params = []
        sql = """
            SELECT c.id, c.code, c.niveau, c.libelle, c.effectif_max, c.actif,
                   COUNT(CASE WHEN i.statut='Inscrit' THEN 1 END) AS effectif
            FROM classes c
            LEFT JOIN inscriptions_eleves i
              ON i.classe_id = c.id
        """
        if self.annee_active:
            sql += " AND i.annee_id = ?"
            params.append(self.annee_active["id"])
        if recherche:
            sql += " WHERE c.code LIKE ? OR c.niveau LIKE ? OR c.libelle LIKE ?"
            terme = f"%{recherche}%"
            params.extend([terme, terme, terme])
        sql += " GROUP BY c.id ORDER BY c.niveau COLLATE NOCASE, c.libelle COLLATE NOCASE"
        c.execute(sql, params)
        lignes = c.fetchall()
        db.close()
        for x in lignes:
            effectif = x["effectif"] or 0
            maximum = x["effectif_max"] if x["effectif_max"] is not None else "—"
            self.table_classes.insert("", "end", iid=str(x["id"]),
                                      values=(x["code"], x["niveau"], x["libelle"], effectif,
                                              maximum, "ACTIVE" if x["actif"] else "INACTIVE"))
        self.classe_selection_info.configure(text=f"{len(lignes)} classe(s) affichée(s). Double-cliquez pour voir les élèves.")

    def obtenir_classe_selectionnee(self):
        selection = self.table_classes.selection()
        if not selection:
            messagebox.showwarning("Aucune sélection", "Veuillez sélectionner une classe.")
            return None
        db = connexion()
        c = db.cursor()
        c.execute("SELECT * FROM classes WHERE id=?", (int(selection[0]),))
        x = c.fetchone()
        db.close()
        return x

    def ouvrir_formulaire_classe(self, classe_id=None):
        f = tk.Toplevel(self)
        f.title("Nouvelle classe" if classe_id is None else "Modifier la classe")
        f.geometry("620x470")
        f.minsize(580, 430)
        f.configure(bg=FOND)
        f.transient(self)
        f.grab_set()
        h = tk.Frame(f, bg=TURQUOISE, height=78)
        h.pack(fill="x")
        h.pack_propagate(False)
        tk.Label(h, text="NOUVELLE CLASSE" if classe_id is None else "MODIFICATION DE LA CLASSE",
                 bg=TURQUOISE, fg=BLANC, font=("Segoe UI", 17, "bold")).pack(anchor="w", padx=25, pady=(16, 0))
        tk.Label(h, text="Définition de la structure pédagogique", bg=TURQUOISE, fg="#D5F5F2",
                 font=("Segoe UI", 9)).pack(anchor="w", padx=25)
        zone = tk.Frame(f, bg=FOND)
        zone.pack(fill="both", expand=True, padx=25, pady=18)
        carte = tk.Frame(zone, bg=CARTE, highlightbackground=BORDURE, highlightthickness=1)
        carte.pack(fill="both", expand=True)
        inter = tk.Frame(carte, bg=CARTE)
        inter.pack(fill="both", expand=True, padx=22, pady=20)
        classe = None
        if classe_id is not None:
            db = connexion()
            c = db.cursor()
            c.execute("SELECT * FROM classes WHERE id=?", (classe_id,))
            classe = c.fetchone()
            db.close()
            if not classe:
                f.destroy()
                messagebox.showerror("Erreur", "La classe demandée n'existe plus.", parent=self)
                return
        champs = {}
        def champ(parent, nom, titre, valeur=""):
            b = tk.Frame(parent, bg=CARTE)
            b.pack(side="left", fill="x", expand=True, padx=5)
            tk.Label(b, text=titre, bg=CARTE, fg=TEXTE, font=("Segoe UI", 9, "bold")).pack(anchor="w", pady=(0, 4))
            e = tk.Entry(b, bg=BLANC, fg=TEXTE, insertbackground=TEXTE, relief="solid", bd=1, font=("Segoe UI", 10))
            e.insert(0, valeur or "")
            e.pack(fill="x", ipady=6)
            champs[nom] = e
        l = tk.Frame(inter, bg=CARTE); l.pack(fill="x", pady=(0, 12))
        champ(l, "code", "Code de la classe *", classe["code"] if classe else "")
        champ(l, "niveau", "Niveau *", classe["niveau"] if classe else "")
        l = tk.Frame(inter, bg=CARTE); l.pack(fill="x", pady=(0, 12))
        champ(l, "libelle", "Nom de la classe *", classe["libelle"] if classe else "")
        champ(l, "effectif_max", "Effectif maximal", classe["effectif_max"] if classe else "")
        actif = tk.BooleanVar(value=bool(classe["actif"]) if classe else True)
        tk.Checkbutton(inter, text="Classe active", variable=actif, bg=CARTE, fg=TEXTE,
                       activebackground=CARTE, selectcolor=CARTE, font=("Segoe UI", 10)).pack(anchor="w", padx=5, pady=5)
        tk.Label(inter, text="Aucun champ Série : la classe est définie par son code, son niveau et son libellé.",
                 bg=CARTE, fg=TEXTE_SECONDAIRE, font=("Segoe UI", 9)).pack(anchor="w", padx=5, pady=(5, 10))
        boutons = tk.Frame(zone, bg=FOND); boutons.pack(fill="x", pady=(12, 0))
        tk.Button(boutons, text="ANNULER", command=f.destroy, bg="#E2E8F0", fg=TEXTE,
                  relief="flat", bd=0, padx=18, pady=9).pack(side="left")
        def enregistrer():
            code = champs["code"].get().strip()
            niveau = champs["niveau"].get().strip()
            libelle = champs["libelle"].get().strip()
            mt = champs["effectif_max"].get().strip()
            maxv = None
            if not code or not niveau or not libelle:
                messagebox.showwarning("Informations manquantes", "Le code, le niveau et le nom de la classe sont obligatoires.", parent=f)
                return
            if mt:
                try:
                    maxv = int(mt)
                    if maxv <= 0:
                        raise ValueError
                except ValueError:
                    messagebox.showwarning("Effectif maximal invalide", "L'effectif maximal doit être un entier positif.", parent=f)
                    return
            db = None
            try:
                db = connexion(); c = db.cursor()
                if classe_id is None:
                    c.execute("INSERT INTO classes(code,niveau,libelle,effectif_max,actif) VALUES(?,?,?,?,?)",
                              (code, niveau, libelle, maxv, 1 if actif.get() else 0))
                else:
                    c.execute("UPDATE classes SET code=?,niveau=?,libelle=?,effectif_max=?,actif=? WHERE id=?",
                              (code, niveau, libelle, maxv, 1 if actif.get() else 0, classe_id))
                db.commit(); db.close(); f.destroy(); self.charger_liste_classes()
                messagebox.showinfo("Enregistrement réussi", "La classe a été enregistrée avec succès.", parent=self)
            except sqlite3.IntegrityError:
                if db: db.close()
                messagebox.showerror("Code déjà utilisé", "Ce code de classe existe déjà.", parent=f)
            except Exception as e:
                if db: db.close()
                messagebox.showerror("Erreur", f"Impossible d'enregistrer la classe.\n\n{e}", parent=f)
        tk.Button(boutons, text="ENREGISTRER", command=enregistrer, bg=TURQUOISE, fg=BLANC,
                  activebackground="#115E59", relief="flat", bd=0, padx=22, pady=10).pack(side="right")
        champs["code"].focus_set()

    def modifier_classe(self):
        x = self.obtenir_classe_selectionnee()
        if x:
            self.ouvrir_formulaire_classe(x["id"])

    def supprimer_classe(self):
        x = self.obtenir_classe_selectionnee()
        if not x:
            return
        db = connexion(); c = db.cursor()
        c.execute("SELECT COUNT(*) AS n FROM inscriptions_eleves WHERE classe_id=?", (x["id"],))
        n = c.fetchone()["n"]
        db.close()
        if n:
            messagebox.showwarning("Suppression impossible", f"Cette classe possède {n} inscription(s).\n\nElle est conservée pour préserver l'historique. Vous pouvez la rendre inactive.")
            return
        if not messagebox.askyesno("Confirmer la suppression", f"Supprimer la classe « {x['libelle']} » ?", icon="warning"):
            return
        db = connexion(); c = db.cursor(); c.execute("DELETE FROM classes WHERE id=?", (x["id"],)); db.commit(); db.close()
        self.charger_liste_classes()
        messagebox.showinfo("Suppression réussie", "La classe a été supprimée.")

    def afficher_eleves_classe_selectionnee(self):
        x = self.obtenir_classe_selectionnee()
        if x:
            self.afficher_liste_eleves_classe(x["id"])

    def afficher_liste_eleves_classe(self, classe_id):
        db = connexion(); c = db.cursor()
        c.execute("SELECT * FROM classes WHERE id=?", (classe_id,)); cl = c.fetchone()
        if not cl:
            db.close(); return
        aid = self.annee_active["id"] if self.annee_active else None
        sql = """SELECT e.matricule,e.nom,e.prenom,e.sexe,e.date_naissance,e.situation_actuelle
                 FROM eleves e JOIN inscriptions_eleves i ON i.eleve_id=e.id
                 WHERE i.classe_id=? AND i.statut='Inscrit'"""
        params = [classe_id]
        if aid:
            sql += " AND i.annee_id=?"; params.append(aid)
        sql += " ORDER BY e.nom COLLATE NOCASE,e.prenom COLLATE NOCASE"
        c.execute(sql, params); eleves = c.fetchall(); db.close()
        self.nettoyer(); self.activer_menu("L'ÉCOLE"); self.titre.configure(text="CLASSES • ÉLÈVES")
        cadre = tk.Frame(self.contenu, bg=FOND); cadre.pack(fill="both", expand=True, padx=25, pady=18)
        en = tk.Frame(cadre, bg=FOND); en.pack(fill="x")
        tk.Label(en, text=f"Élèves de {cl['libelle']}", bg=FOND, fg=TEXTE, font=("Segoe UI", 20, "bold")).pack(side="left")
        tk.Button(en, text="← RETOUR AUX CLASSES", command=self.module_classes, bg="#E2E8F0", fg=TEXTE,
                  relief="flat", bd=0, padx=15, pady=9).pack(side="right")
        effectif = len(eleves)
        maximum = f" / {cl['effectif_max']}" if cl["effectif_max"] else ""
        tk.Label(cadre, text=f"Code : {cl['code']}  •  Niveau : {cl['niveau']}  •  Effectif : {effectif}{maximum}",
                 bg=FOND, fg=TEXTE_SECONDAIRE, font=("Segoe UI", 10)).pack(anchor="w", pady=(3, 15))
        zone = tk.Frame(cadre, bg=CARTE, highlightbackground=BORDURE, highlightthickness=1); zone.pack(fill="both", expand=True)
        cols = ("matricule","nom","prenom","sexe","date","situation")
        t = ttk.Treeview(zone, columns=cols, show="headings")
        heads = {"matricule":"MATRICULE","nom":"NOM","prenom":"PRÉNOM(S)","sexe":"SEXE","date":"DATE DE NAISSANCE","situation":"SITUATION"}
        widths = {"matricule":120,"nom":160,"prenom":180,"sexe":80,"date":130,"situation":120}
        for col in cols:
            t.heading(col, text=heads[col]); t.column(col, width=widths[col], minwidth=75, anchor="w")
        for e in eleves:
            t.insert("", "end", values=(e["matricule"],e["nom"],e["prenom"],e["sexe"],e["date_naissance"] or "",e["situation_actuelle"]))
        sy = ttk.Scrollbar(zone, orient="vertical", command=t.yview); sy.pack(side="right", fill="y"); t.configure(yscrollcommand=sy.set); t.pack(side="left", fill="both", expand=True)

    def _val_excel(self, row, idx, key):
        i = idx.get(key)
        if i is None or i >= len(row) or row[i] is None:
            return ""
        value = row[i]
        if hasattr(value, "strftime") and key == "date_naissance":
            return value.strftime("%d/%m/%Y")
        return str(value).strip()

    def importer_eleves_classe_excel(self):
        cl = self.obtenir_classe_selectionnee()
        if not cl: return
        if not self.annee_active:
            messagebox.showwarning("Année scolaire", "Aucune année scolaire active n'est définie.")
            return
        if load_workbook is None:
            messagebox.showerror("Module manquant", "openpyxl n'est pas installé.\n\nCommande : python -m pip install openpyxl")
            return
        fichier = filedialog.askopenfilename(title=f"Importer les élèves dans {cl['libelle']}",
                                             filetypes=[("Fichiers Excel", "*.xlsx"), ("Tous les fichiers", "*.*")])
        if not fichier: return
        try:
            wb = load_workbook(fichier, read_only=True, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            wb.close()
            if not rows:
                messagebox.showwarning("Fichier vide", "Le fichier Excel ne contient aucune donnée.")
                return
            headers = [str(v).strip().lower() if v is not None else "" for v in rows[0]]
            aliases = {
                "matricule":{"matricule","matricule élève","matricule eleve","id élève","id eleve"},
                "nom":{"nom","nom de famille"},
                "prenom":{"prenom","prénom","prénom(s)","prenom(s)","prenoms","prénoms"},
                "sexe":{"sexe","genre"},
                "date_naissance":{"date de naissance","date_naissance","date naissance"},
                "lieu_naissance":{"lieu de naissance","lieu_naissance"},
                "nationalite":{"nationalité","nationalite"},
                "statut_entree":{"statut d'entrée","statut entree","statut_entree","statut"},
                "situation_actuelle":{"situation actuelle","situation_actuelle","situation"},
                "nom_parent":{"parent","responsable","nom parent","nom_parent"},
                "telephone_parent":{"téléphone parent","telephone parent","téléphone responsable","telephone responsable","telephone_parent"},
                "observation":{"observation","observations"}
            }
            idx = {}
            for key, names in aliases.items():
                for i, h in enumerate(headers):
                    if h in names:
                        idx[key] = i; break
            missing = [x for x in ("matricule","nom","prenom") if x not in idx]
            if missing:
                messagebox.showerror("Colonnes manquantes", "Le fichier doit contenir au minimum : MATRICULE, NOM et PRÉNOM(S).")
                return

            db = connexion(); c = db.cursor()
            aj = maj = ins = deja = limites = 0
            erreurs = []
            # Comptage de la classe pour l'année active.
            c.execute("SELECT COUNT(*) AS n FROM inscriptions_eleves WHERE classe_id=? AND annee_id=? AND statut='Inscrit'",
                      (cl["id"], self.annee_active["id"]))
            effectif_initial = c.fetchone()["n"] or 0
            nouveaux_dans_classe = 0
            for no, row in enumerate(rows[1:], 2):
                matricule = self._val_excel(row, idx, "matricule")
                nom = self._val_excel(row, idx, "nom")
                prenom = self._val_excel(row, idx, "prenom")
                if not matricule and not nom and not prenom: continue
                if not matricule or not nom or not prenom:
                    erreurs.append(f"Ligne {no} : matricule, nom ou prénom manquant."); continue
                try:
                    sexe_brut = self._val_excel(row, idx, "sexe").lower()
                    if sexe_brut in ("m","masculin","garçon","garcon","male","h"):
                        sexe = "Masculin"
                    elif sexe_brut in ("f","féminin","feminin","fille","female"):
                        sexe = "Féminin"
                    else:
                        sexe = "Féminin"
                    vals = (nom, prenom, sexe, self._val_excel(row, idx, "date_naissance"),
                            self._val_excel(row, idx, "lieu_naissance"), self._val_excel(row, idx, "nationalite"),
                            self._val_excel(row, idx, "statut_entree") or "Nouveau",
                            self._val_excel(row, idx, "situation_actuelle") or "Actif",
                            self._val_excel(row, idx, "nom_parent"), self._val_excel(row, idx, "telephone_parent"),
                            self._val_excel(row, idx, "observation"))
                    c.execute("SELECT id FROM eleves WHERE matricule=?", (matricule,)); e = c.fetchone()
                    if e:
                        eid = e["id"]
                        c.execute("UPDATE eleves SET nom=?,prenom=?,sexe=?,date_naissance=?,lieu_naissance=?,nationalite=?,statut_entree=?,situation_actuelle=?,nom_parent=?,telephone_parent=?,observation=? WHERE id=?", vals + (eid,))
                        maj += 1
                    else:
                        c.execute("INSERT INTO eleves(matricule,nom,prenom,sexe,date_naissance,lieu_naissance,nationalite,statut_entree,situation_actuelle,nom_parent,telephone_parent,observation) VALUES(?,?,?,?,?,?,?,?,?,?,?,?)", (matricule,) + vals)
                        eid = c.lastrowid; aj += 1
                    c.execute("SELECT id,classe_id,statut FROM inscriptions_eleves WHERE eleve_id=? AND annee_id=?", (eid, self.annee_active["id"]))
                    inscription = c.fetchone()
                    if inscription:
                        if inscription["classe_id"] == cl["id"] and inscription["statut"] == "Inscrit":
                            deja += 1
                        else:
                            if cl["effectif_max"] and effectif_initial + nouveaux_dans_classe >= cl["effectif_max"]:
                                limites += 1
                                erreurs.append(f"Ligne {no} : effectif maximal de la classe atteint.")
                                continue
                            c.execute("UPDATE inscriptions_eleves SET classe_id=?, statut='Inscrit' WHERE id=?", (cl["id"], inscription["id"]))
                            ins += 1; nouveaux_dans_classe += 1
                    else:
                        if cl["effectif_max"] and effectif_initial + nouveaux_dans_classe >= cl["effectif_max"]:
                            limites += 1
                            erreurs.append(f"Ligne {no} : effectif maximal de la classe atteint.")
                            continue
                        c.execute("INSERT INTO inscriptions_eleves(eleve_id,annee_id,classe_id,statut) VALUES(?,?,?,'Inscrit')", (eid,self.annee_active["id"],cl["id"]))
                        ins += 1; nouveaux_dans_classe += 1
                except Exception as erreur:
                    erreurs.append(f"Ligne {no} : {erreur}")
            db.commit(); db.close()
            msg = (f"Import terminé pour {cl['libelle']}.\n\nNouveaux dossiers : {aj}\n"
                   f"Dossiers mis à jour : {maj}\nInscriptions créées ou déplacées : {ins}\n"
                   f"Déjà présents : {deja}")
            if limites: msg += f"\nLimites d'effectif : {limites}"
            if erreurs: msg += f"\n\nLignes à vérifier : {len(erreurs)}\n" + "\n".join(erreurs[:10])
            messagebox.showinfo("Import Excel terminé", msg)
            self.afficher_liste_eleves_classe(cl["id"])
        except Exception as e:
            messagebox.showerror("Erreur d'importation", f"Impossible de lire le fichier Excel.\n\n{e}")

    def creer_modele_excel_classe(self):
        if Workbook is None:
            messagebox.showerror("Module manquant", "openpyxl n'est pas installé.\n\nCommande : python -m pip install openpyxl")
            return
        fichier = filedialog.asksaveasfilename(title="Enregistrer le modèle Excel des élèves",
                                               defaultextension=".xlsx",
                                               filetypes=[("Fichier Excel", "*.xlsx")],
                                               initialfile="modele_eleves_classe.xlsx")
        if not fichier: return
        try:
            wb = Workbook(); ws = wb.active; ws.title = "Élèves"
            headers = ["MATRICULE","NOM","PRÉNOM(S)","SEXE","DATE DE NAISSANCE","LIEU DE NAISSANCE","NATIONALITÉ","STATUT D'ENTRÉE","SITUATION ACTUELLE","PARENT / RESPONSABLE","TÉLÉPHONE","OBSERVATION"]
            ws.append(headers)
            ws.append(["EX001","DOE","Jean","Masculin","15/09/2012","Lomé","Togolaise","Nouveau","Actif","Nom du responsable","90 00 00 00",""])
            ws.freeze_panes = "A2"
            for cell in ws[1]: cell.font = cell.font.copy(bold=True)
            largeurs = [18,20,25,14,20,22,18,20,22,28,20,35]
            for i, largeur in enumerate(largeurs, 1): ws.column_dimensions[chr(64+i)].width = largeur
            wb.save(fichier)
            messagebox.showinfo("Modèle créé", "Le modèle Excel a été créé avec succès.")
        except Exception as e:
            messagebox.showerror("Erreur", f"Impossible de créer le modèle Excel.\n\n{e}")

    def exporter_eleves_classe_excel(self):
        cl = self.obtenir_classe_selectionnee()
        if not cl: return
        if not self.annee_active:
            messagebox.showwarning("Année scolaire", "Aucune année scolaire active n'est définie.")
            return
        if Workbook is None:
            messagebox.showerror("Module manquant", "openpyxl n'est pas installé.\n\nCommande : python -m pip install openpyxl")
            return
        db = connexion(); c = db.cursor()
        c.execute("""SELECT e.matricule,e.nom,e.prenom,e.sexe,e.date_naissance,e.lieu_naissance,e.nationalite,
                           e.statut_entree,e.situation_actuelle,e.nom_parent,e.telephone_parent,e.observation
                    FROM eleves e JOIN inscriptions_eleves i ON i.eleve_id=e.id
                    WHERE i.classe_id=? AND i.annee_id=? AND i.statut='Inscrit'
                    ORDER BY e.nom COLLATE NOCASE,e.prenom COLLATE NOCASE""", (cl["id"], self.annee_active["id"]))
        eleves = c.fetchall(); db.close()
        fichier = filedialog.asksaveasfilename(title=f"Exporter les élèves de {cl['libelle']}",
                                               defaultextension=".xlsx",
                                               filetypes=[("Fichier Excel", "*.xlsx")],
                                               initialfile=f"eleves_{cl['code']}.xlsx")
        if not fichier: return
        try:
            wb = Workbook(); ws = wb.active; ws.title = cl["libelle"][:31]
            headers = ["MATRICULE","NOM","PRÉNOM(S)","SEXE","DATE DE NAISSANCE","LIEU DE NAISSANCE","NATIONALITÉ","STATUT D'ENTRÉE","SITUATION ACTUELLE","PARENT / RESPONSABLE","TÉLÉPHONE","OBSERVATION"]
            ws.append(headers)
            for e in eleves:
                ws.append([e["matricule"],e["nom"],e["prenom"],e["sexe"],e["date_naissance"] or "",e["lieu_naissance"] or "",e["nationalite"] or "",e["statut_entree"] or "",e["situation_actuelle"] or "",e["nom_parent"] or "",e["telephone_parent"] or "",e["observation"] or ""])
            ws.freeze_panes = "A2"
            for cell in ws[1]: cell.font = cell.font.copy(bold=True)
            for i, largeur in enumerate([18,20,25,14,20,22,18,20,22,28,20,35], 1): ws.column_dimensions[chr(64+i)].width = largeur
            wb.save(fichier)
            messagebox.showinfo("Export Excel terminé", f"{len(eleves)} élève(s) ont été exportés avec succès.")
        except Exception as e:
            messagebox.showerror("Erreur", f"Impossible d'exporter la liste Excel.\n\n{e}")

    # ========================================================
    # MODULE MATIÈRES
    # ========================================================

    def module_matieres(self):

        self.afficher_page(
            "MATIÈRES",
            "Gestion des matières et des coefficients.",
            "Le module de gestion des matières sera construit ici.",
            "L'ÉCOLE"
        )


    # ========================================================
    # EMPLOI DU TEMPS
    # ========================================================

    def afficher_emploi(self):

        self.afficher_page(
            "EMPLOI DU TEMPS",
            "Organisation des cours et des horaires.",
            "Le module Emploi du temps sera construit ici.",
            "EMPLOI DU TEMPS"
        )


    # ========================================================
    # CONSEIL
    # ========================================================

    def afficher_conseil(self):

        self.afficher_page(
            "CONSEIL",
            "Gestion des conseils de classe.",
            "Le module Conseil sera construit ici.",
            "CONSEIL"
        )


    # ========================================================
    # DÉVELOPPEUR
    # ========================================================

    def afficher_developpeur(self):

        self.nettoyer()

        self.activer_menu("DÉVELOPPEUR")

        self.titre.configure(
            text="DÉVELOPPEUR"
        )

        cadre = tk.Frame(
            self.contenu,
            bg=FOND
        )

        cadre.pack(
            fill="both",
            expand=True,
            padx=25,
            pady=22
        )

        tk.Label(
            cadre,
            text="DÉVELOPPEUR",
            bg=FOND,
            fg=TEXTE,
            font=("Segoe UI", 22, "bold")
        ).pack(anchor="w")

        tk.Label(
            cadre,
            text="Administration technique de GestionScolaire",
            bg=FOND,
            fg=TEXTE_SECONDAIRE,
            font=("Segoe UI", 10)
        ).pack(
            anchor="w",
            pady=(3, 20)
        )

        grille = tk.Frame(
            cadre,
            bg=FOND
        )

        grille.pack(
            fill="both",
            expand=True
        )

        # ----------------------------------------------------
        # INFORMATIONS DE L'ÉCOLE
        # ----------------------------------------------------

        self.creer_module(
            grille,
            "INFORMATIONS DE L'ÉCOLE",
            "Saisir ou modifier les informations de l'école",
            TURQUOISE,
            0,
            0,
            self.afficher_parametres
        )

        # ----------------------------------------------------
        # SAUVEGARDE
        # ----------------------------------------------------

        self.creer_module(
            grille,
            "SAUVEGARDE",
            "Gestion des sauvegardes locales",
            VERT,
            0,
            1,
            self.module_sauvegarde
        )

        # ----------------------------------------------------
        # MAINTENANCE
        # ----------------------------------------------------

        self.creer_module(
            grille,
            "MAINTENANCE",
            "Outils de maintenance du logiciel",
            ORANGE,
            1,
            0,
            self.module_maintenance
        )

        # ----------------------------------------------------
        # INFORMATIONS SYSTÈME
        # ----------------------------------------------------

        self.creer_module(
            grille,
            "INFORMATIONS SYSTÈME",
            "Informations sur GestionScolaire",
            BLEU,
            1,
            1,
            self.module_systeme
        )


    # ========================================================
    # MODULE SAUVEGARDE
    # ========================================================

    def module_sauvegarde(self):

        self.afficher_page(
            "SAUVEGARDE",
            "Gestion des sauvegardes locales.",
            "Le module de sauvegarde sera construit ici.",
            "DÉVELOPPEUR"
        )


    # ========================================================
    # MODULE MAINTENANCE
    # ========================================================

    def module_maintenance(self):

        self.afficher_page(
            "MAINTENANCE",
            "Outils de maintenance du logiciel.",
            "Le module de maintenance sera construit ici.",
            "DÉVELOPPEUR"
        )


    # ========================================================
    # MODULE INFORMATIONS SYSTÈME
    # ========================================================

    def module_systeme(self):

        self.afficher_page(
            "INFORMATIONS SYSTÈME",
            "Informations techniques sur GestionScolaire.",
            "Version locale de GestionScolaire.",
            "DÉVELOPPEUR"
        )


    # ========================================================
    # PARAMÈTRES DE L'ÉCOLE
    # ========================================================

    def afficher_parametres(self):

        self.nettoyer()

        self.activer_menu("DÉVELOPPEUR")

        self.titre.configure(
            text="INFORMATIONS DE L'ÉCOLE"
        )

        # ----------------------------------------------------
        # CONTENEUR PRINCIPAL
        # ----------------------------------------------------

        cadre = tk.Frame(
            self.contenu,
            bg=FOND
        )

        cadre.pack(
            fill="both",
            expand=True,
            padx=25,
            pady=22
        )

        tk.Label(
            cadre,
            text="Informations de l'école",
            bg=FOND,
            fg=TEXTE,
            font=("Segoe UI", 22, "bold")
        ).pack(anchor="w")

        tk.Label(
            cadre,
            text=(
                "Cette page permet de saisir ou modifier "
                "les informations générales de l'établissement."
            ),
            bg=FOND,
            fg=TEXTE_SECONDAIRE,
            font=("Segoe UI", 10)
        ).pack(
            anchor="w",
            pady=(3, 12)
        )

        # ----------------------------------------------------
        # ZONE AVEC DÉFILEMENT
        # ----------------------------------------------------

        zone = tk.Frame(
            cadre,
            bg=CARTE,
            highlightbackground=BORDURE,
            highlightthickness=1
        )

        zone.pack(
            fill="both",
            expand=True
        )

        canvas = tk.Canvas(
            zone,
            bg=CARTE,
            highlightthickness=0,
            bd=0
        )

        scrollbar = tk.Scrollbar(
            zone,
            orient="vertical",
            command=canvas.yview
        )

        canvas.configure(
            yscrollcommand=scrollbar.set
        )

        scrollbar.pack(
            side="right",
            fill="y"
        )

        canvas.pack(
            side="left",
            fill="both",
            expand=True
        )

        contenu = tk.Frame(
            canvas,
            bg=CARTE
        )

        fenetre_canvas = canvas.create_window(
            (0, 0),
            window=contenu,
            anchor="nw"
        )

        def mettre_a_jour_scroll(event=None):
            canvas.configure(
                scrollregion=canvas.bbox("all")
            )

        def ajuster_largeur(event):
            canvas.itemconfigure(
                fenetre_canvas,
                width=event.width
            )

        contenu.bind(
            "<Configure>",
            mettre_a_jour_scroll
        )

        canvas.bind(
            "<Configure>",
            ajuster_largeur
        )

        # ----------------------------------------------------
        # DÉFILEMENT À LA MOLETTE
        # ----------------------------------------------------

        def roulette(event):

            canvas.yview_scroll(
                int(-1 * (event.delta / 120)),
                "units"
            )

        canvas.bind_all(
            "<MouseWheel>",
            roulette
        )

        # ----------------------------------------------------
        # CONTENU
        # ----------------------------------------------------

        interieur = tk.Frame(
            contenu,
            bg=CARTE
        )

        interieur.pack(
            fill="both",
            expand=True,
            padx=30,
            pady=22
        )

        # ----------------------------------------------------
        # NOM
        # ----------------------------------------------------

        self.param_nom = self.creer_champ(
            interieur,
            "Nom de l'établissement *",
            self.etablissement["nom"] or ""
        )

        # ----------------------------------------------------
        # TYPE
        # ----------------------------------------------------

        tk.Label(
            interieur,
            text="Type d'établissement",
            bg=CARTE,
            fg=TEXTE,
            font=("Segoe UI", 9, "bold")
        ).pack(
            anchor="w",
            pady=(12, 4)
        )

        self.param_type = tk.StringVar(
            value=(
                self.etablissement["type_etablissement"]
                or "CEG"
            )
        )

        menu_type = tk.OptionMenu(
            interieur,
            self.param_type,
            "CEG",
            "Collège",
            "Lycée",
            "Collège-Lycée",
            "École primaire",
            "Établissement privé",
            "Autre"
        )

        menu_type.configure(
            bg=BLANC,
            fg=TEXTE,
            activebackground=BLANC,
            activeforeground=TEXTE,
            relief="solid",
            bd=1,
            highlightthickness=0,
            font=("Segoe UI", 10)
        )

        menu_type.pack(
            fill="x"
        )

        # ----------------------------------------------------
        # LIGNE 1
        # ----------------------------------------------------

        ligne1 = tk.Frame(
            interieur,
            bg=CARTE
        )

        ligne1.pack(
            fill="x",
            pady=(12, 0)
        )

        gauche1 = tk.Frame(
            ligne1,
            bg=CARTE
        )

        gauche1.pack(
            side="left",
            fill="x",
            expand=True,
            padx=(0, 8)
        )

        droite1 = tk.Frame(
            ligne1,
            bg=CARTE
        )

        droite1.pack(
            side="left",
            fill="x",
            expand=True,
            padx=(8, 0)
        )

        self.param_adresse = self.creer_champ(
            gauche1,
            "Adresse",
            self.etablissement["adresse"] or ""
        )

        self.param_localite = self.creer_champ(
            droite1,
            "Localité",
            self.etablissement["localite"] or ""
        )

        # ----------------------------------------------------
        # LIGNE 2
        # ----------------------------------------------------

        ligne2 = tk.Frame(
            interieur,
            bg=CARTE
        )

        ligne2.pack(
            fill="x",
            pady=(12, 0)
        )

        gauche2 = tk.Frame(
            ligne2,
            bg=CARTE
        )

        gauche2.pack(
            side="left",
            fill="x",
            expand=True,
            padx=(0, 8)
        )

        droite2 = tk.Frame(
            ligne2,
            bg=CARTE
        )

        droite2.pack(
            side="left",
            fill="x",
            expand=True,
            padx=(8, 0)
        )

        self.param_telephone = self.creer_champ(
            gauche2,
            "Téléphone",
            self.etablissement["telephone"] or ""
        )

        self.param_email = self.creer_champ(
            droite2,
            "E-mail",
            self.etablissement["email"] or ""
        )

        # ----------------------------------------------------
        # LIGNE 3
        # ----------------------------------------------------

        ligne3 = tk.Frame(
            interieur,
            bg=CARTE
        )

        ligne3.pack(
            fill="x",
            pady=(12, 0)
        )

        gauche3 = tk.Frame(
            ligne3,
            bg=CARTE
        )

        gauche3.pack(
            side="left",
            fill="x",
            expand=True,
            padx=(0, 8)
        )

        droite3 = tk.Frame(
            ligne3,
            bg=CARTE
        )

        droite3.pack(
            side="left",
            fill="x",
            expand=True,
            padx=(8, 0)
        )

        self.param_directeur = self.creer_champ(
            gauche3,
            "Directeur / Chef d'établissement",
            self.etablissement["directeur"] or ""
        )

        self.param_secretaire = self.creer_champ(
            droite3,
            "Secrétaire",
            self.etablissement["secretaire"] or ""
        )

        # ----------------------------------------------------
        # LIGNE 4
        # ----------------------------------------------------

        ligne4 = tk.Frame(
            interieur,
            bg=CARTE
        )

        ligne4.pack(
            fill="x",
            pady=(12, 0)
        )

        gauche4 = tk.Frame(
            ligne4,
            bg=CARTE
        )

        gauche4.pack(
            side="left",
            fill="x",
            expand=True,
            padx=(0, 8)
        )

        droite4 = tk.Frame(
            ligne4,
            bg=CARTE
        )

        droite4.pack(
            side="left",
            fill="x",
            expand=True,
            padx=(8, 0)
        )

        self.param_dre = self.creer_champ(
            gauche4,
            "DRE / Inspection",
            self.etablissement["dre_inspection"] or ""
        )

        annee = (
            self.annee_active["libelle"]
            if self.annee_active
            else ""
        )

        self.param_annee = self.creer_champ(
            droite4,
            "Année scolaire active *",
            annee
        )

        # ----------------------------------------------------
        # DEVISE
        # ----------------------------------------------------

        self.param_devise = self.creer_champ(
            interieur,
            "Devise / Slogan de l'établissement",
            self.etablissement["devise"] or ""
        )

        # ----------------------------------------------------
        # LOGO
        # ----------------------------------------------------

        bloc_logo = tk.Frame(
            interieur,
            bg=CARTE
        )

        bloc_logo.pack(
            fill="x",
            pady=(12, 0)
        )

        tk.Label(
            bloc_logo,
            text="Logo de l'établissement",
            bg=CARTE,
            fg=TEXTE,
            font=("Segoe UI", 9, "bold")
        ).pack(
            anchor="w",
            pady=(0, 4)
        )

        ligne_logo = tk.Frame(
            bloc_logo,
            bg=CARTE
        )

        ligne_logo.pack(
            fill="x"
        )

        ancien_logo = (
            self.etablissement["logo"]
            or ""
        )

        nom_logo = (
            Path(ancien_logo).name
            if ancien_logo
            else "Aucun logo enregistré"
        )

        self.param_logo_texte = tk.StringVar(
            value=nom_logo
        )

        tk.Label(
            ligne_logo,
            textvariable=self.param_logo_texte,
            bg=BLANC,
            fg=TEXTE_SECONDAIRE,
            anchor="w",
            relief="solid",
            bd=1,
            padx=10,
            font=("Segoe UI", 9)
        ).pack(
            side="left",
            fill="x",
            expand=True,
            ipady=6
        )

        tk.Button(
            ligne_logo,
            text="CHOISIR LE LOGO",
            command=self.choisir_logo_parametre,
            bg=BLEU,
            fg=BLANC,
            activebackground="#2563EB",
            activeforeground=BLANC,
            relief="flat",
            bd=0,
            cursor="hand2",
            font=("Segoe UI", 9, "bold"),
            padx=15,
            pady=7
        ).pack(
            side="left",
            padx=(8, 0)
        )

        # ----------------------------------------------------
        # BOUTONS
        # ----------------------------------------------------

        boutons = tk.Frame(
            interieur,
            bg=CARTE
        )

        boutons.pack(
            fill="x",
            pady=(22, 10)
        )

        tk.Button(
            boutons,
            text="← RETOUR",
            command=self.afficher_developpeur,
            bg="#E2E8F0",
            fg=TEXTE,
            activebackground="#CBD5E1",
            activeforeground=TEXTE,
            relief="flat",
            bd=0,
            cursor="hand2",
            font=("Segoe UI", 9, "bold"),
            padx=18,
            pady=10
        ).pack(
            side="left"
        )

        tk.Button(
            boutons,
            text="ENREGISTRER LES MODIFICATIONS",
            command=self.enregistrer_parametres,
            bg=TURQUOISE,
            fg=BLANC,
            activebackground="#115E59",
            activeforeground=BLANC,
            relief="flat",
            bd=0,
            cursor="hand2",
            font=("Segoe UI", 10, "bold"),
            padx=22,
            pady=11
        ).pack(
            side="right"
        )

        # ----------------------------------------------------
        # RETOUR EN HAUT
        # ----------------------------------------------------

        canvas.yview_moveto(0)

    # ========================================================
    # CHOISIR LOGO PARAMÈTRES
    # ========================================================

    def choisir_logo_parametre(self):

        fichier = filedialog.askopenfilename(
            title="Choisir le logo de l'établissement",
            filetypes=[
                (
                    "Images",
                    "*.png *.jpg *.jpeg *.gif *.bmp"
                ),
                ("PNG", "*.png"),
                ("JPEG", "*.jpg *.jpeg"),
                ("Tous les fichiers", "*.*")
            ]
        )

        if not fichier:
            return

        try:

            source = Path(fichier)

            destination = DOSSIER_LOGO / (
                "logo_etablissement"
                + source.suffix.lower()
            )

            shutil.copy2(
                source,
                destination
            )

            self.logo_selectionne = str(destination)

            self.param_logo_texte.set(
                source.name
            )

        except Exception as erreur:

            messagebox.showerror(
                "Erreur",
                f"Impossible de copier le logo.\n\n{erreur}"
            )


    # ========================================================
    # ENREGISTRER LES PARAMÈTRES
    # ========================================================

    def enregistrer_parametres(self):

        nom = self.param_nom.get().strip()

        annee = self.param_annee.get().strip()

        if not nom:

            messagebox.showwarning(
                "Information manquante",
                "Le nom de l'établissement est obligatoire."
            )

            return

        if not annee:

            messagebox.showwarning(
                "Information manquante",
                "L'année scolaire est obligatoire."
            )

            return

        logo = self.logo_selectionne

        if not logo:

            logo = (
                self.etablissement["logo"]
                or ""
            )

        try:

            enregistrer_ecole(
                nom,
                self.param_type.get(),
                self.param_adresse.get().strip(),
                self.param_localite.get().strip(),
                self.param_telephone.get().strip(),
                self.param_email.get().strip(),
                self.param_dre.get().strip(),
                self.param_directeur.get().strip(),
                self.param_secretaire.get().strip(),
                self.param_devise.get().strip(),
                logo,
                annee
            )

            self.charger_donnees()

            self.logo_selectionne = ""

            messagebox.showinfo(
                "Enregistrement réussi",
                "Les informations de l'école "
                "ont été enregistrées avec succès."
            )

            self.construire_interface_principale()

        except Exception as erreur:

            messagebox.showerror(
                "Erreur",
                f"Impossible d'enregistrer les modifications.\n\n{erreur}"
            )


    # ========================================================
    # PAGE GÉNÉRIQUE
    # ========================================================

    def afficher_page(
        self,
        titre,
        sous_titre,
        message,
        menu
    ):

        self.nettoyer()

        self.activer_menu(menu)

        self.titre.configure(
            text=titre
        )

        cadre = tk.Frame(
            self.contenu,
            bg=FOND
        )

        cadre.pack(
            fill="both",
            expand=True,
            padx=25,
            pady=22
        )

        tk.Label(
            cadre,
            text=titre,
            bg=FOND,
            fg=TEXTE,
            font=("Segoe UI", 22, "bold")
        ).pack(anchor="w")

        tk.Label(
            cadre,
            text=sous_titre,
            bg=FOND,
            fg=TEXTE_SECONDAIRE,
            font=("Segoe UI", 10)
        ).pack(
            anchor="w",
            pady=(3, 20)
        )

        bloc = tk.Frame(
            cadre,
            bg=CARTE,
            highlightbackground=BORDURE,
            highlightthickness=1
        )

        bloc.pack(
            fill="both",
            expand=True
        )

        tk.Label(
            bloc,
            text=message,
            bg=CARTE,
            fg=TEXTE_SECONDAIRE,
            font=("Segoe UI", 11)
        ).pack(
            pady=60
        )


    # ========================================================
    # CARTES TABLEAU DE BORD
    # ========================================================

    def creer_carte(
        self,
        parent,
        titre,
        valeur,
        couleur,
        description
    ):

        carte = tk.Frame(
            parent,
            bg=CARTE,
            highlightbackground=BORDURE,
            highlightthickness=1
        )

        carte.pack(
            side="left",
            fill="both",
            expand=True,
            padx=5
        )

        tk.Frame(
            carte,
            bg=couleur,
            width=5
        ).pack(
            side="left",
            fill="y"
        )

        interieur = tk.Frame(
            carte,
            bg=CARTE
        )

        interieur.pack(
            side="left",
            fill="both",
            expand=True,
            padx=15,
            pady=12
        )

        tk.Label(
            interieur,
            text=titre,
            bg=CARTE,
            fg=TEXTE_SECONDAIRE,
            font=("Segoe UI", 9, "bold")
        ).pack(anchor="w")

        tk.Label(
            interieur,
            text=valeur,
            bg=CARTE,
            fg=TEXTE,
            font=("Segoe UI", 25, "bold")
        ).pack(anchor="w")

        tk.Label(
            interieur,
            text=description,
            bg=CARTE,
            fg=TEXTE_SECONDAIRE,
            font=("Segoe UI", 8)
        ).pack(anchor="w")


    # ========================================================
    # MODULES
    # ========================================================

    def creer_module(
        self,
        parent,
        titre,
        description,
        couleur,
        ligne,
        colonne,
        commande=None
    ):

        module = tk.Frame(
            parent,
            bg=CARTE,
            highlightbackground=BORDURE,
            highlightthickness=1,
            cursor="hand2" if commande else ""
        )

        module.grid(
            row=ligne,
            column=colonne,
            sticky="nsew",
            padx=7,
            pady=7
        )

        parent.grid_rowconfigure(
            ligne,
            weight=1
        )

        parent.grid_columnconfigure(
            colonne,
            weight=1
        )

        barre_couleur = tk.Frame(
            module,
            bg=couleur,
            width=6,
            cursor="hand2" if commande else ""
        )

        barre_couleur.pack(
            side="left",
            fill="y"
        )

        contenu = tk.Frame(
            module,
            bg=CARTE,
            cursor="hand2" if commande else ""
        )

        contenu.pack(
            fill="both",
            expand=True,
            padx=20,
            pady=20
        )

        label_titre = tk.Label(
            contenu,
            text=titre,
            bg=CARTE,
            fg=TEXTE,
            font=("Segoe UI", 14, "bold"),
            cursor="hand2" if commande else ""
        )

        label_titre.pack(
            anchor="w"
        )

        label_description = tk.Label(
            contenu,
            text=description,
            bg=CARTE,
            fg=TEXTE_SECONDAIRE,
            font=("Segoe UI", 10),
            cursor="hand2" if commande else ""
        )

        label_description.pack(
            anchor="w",
            pady=(7, 0)
        )

        # ----------------------------------------------------
        # CORRECTION IMPORTANTE :
        # LE CLIC EST ATTACHÉ À TOUS LES ÉLÉMENTS DU MODULE
        # ----------------------------------------------------

        if commande:

            widgets_cliquables = (
                module,
                barre_couleur,
                contenu,
                label_titre,
                label_description
            )

            for widget in widgets_cliquables:

                widget.bind(
                    "<Button-1>",
                    lambda event, f=commande: f()
                )

                widget.bind(
                    "<Enter>",
                    lambda event,
                    m=module,
                    b=barre_couleur,
                    c=contenu,
                    t=label_titre,
                    d=label_description:
                    self.module_survol(
                        m,
                        b,
                        c,
                        t,
                        d
                    )
                )

                widget.bind(
                    "<Leave>",
                    lambda event,
                    m=module,
                    b=barre_couleur,
                    c=contenu,
                    t=label_titre,
                    d=label_description:
                    self.module_quitte(
                        m,
                        b,
                        c,
                        t,
                        d
                    )
                )


    # ========================================================
    # SURVOL MODULE
    # ========================================================

    def module_survol(
        self,
        module,
        barre,
        contenu,
        titre,
        description
    ):

        module.configure(
            bg="#F1F5F9"
        )

        contenu.configure(
            bg="#F1F5F9"
        )

        titre.configure(
            bg="#F1F5F9"
        )

        description.configure(
            bg="#F1F5F9"
        )


    # ========================================================
    # FIN SURVOL MODULE
    # ========================================================

    def module_quitte(
        self,
        module,
        barre,
        contenu,
        titre,
        description
    ):

        module.configure(
            bg=CARTE
        )

        contenu.configure(
            bg=CARTE
        )

        titre.configure(
            bg=CARTE
        )

        description.configure(
            bg=CARTE
        )


# ============================================================
# LANCEMENT
# ============================================================

if __name__ == "__main__":

    application = GestionScolaire()

    application.mainloop()